#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel写入模块
- 一企业一文件
- 每年度每个财务报表一个sheet
- 列名均为中文
"""

import os
import logging
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from config import config

logger = logging.getLogger(__name__)


# Excel样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 数值列标识（根据字段名判断）
NUMERIC_KEYWORDS = [
    "金额", "合计", "总计", "净额", "净收入", "净支出", "净利润",
    "总额", "余额", "小计", "收益", "损益", "支出", "收入",
    "成本", "费用", "利润", "资产", "负债", "权益", "税",
    "现金", "借款", "股本", "资本", "盈余", "股", "利息",
    "工资", "每股", "折旧", "摊销",
]

# 人民币数字格式
CNY_FORMAT = '#,##0.00'


def _is_numeric_column(header: str) -> bool:
    """判断列是否为数值类型"""
    for kw in NUMERIC_KEYWORDS:
        if kw in header:
            return True
    return False


def _format_sheet(ws, headers: List[str], num_cols: int, num_rows: int):
    """
    美化Excel工作表格式

    参数:
        ws: 工作表对象
        headers: 表头列表
        num_cols: 列数
        num_rows: 数据行数（不含表头）
    """
    # 设置列宽
    for col_idx in range(1, num_cols + 1):
        header_text = headers[col_idx - 1] if col_idx <= len(headers) else ""
        if len(header_text) > 10:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(header_text) * 2.2)
        elif _is_numeric_column(header_text):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # 设置表头样式
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 设置数据行样式
    for row_idx in range(2, num_rows + 2):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            header = headers[col_idx - 1] if col_idx <= len(headers) else ""
            if _is_numeric_column(header):
                cell.alignment = NUMBER_ALIGNMENT
                if isinstance(cell.value, (int, float)):
                    cell.number_format = CNY_FORMAT
            else:
                cell.alignment = DATA_ALIGNMENT
                cell.font = DATA_FONT

    # 冻结首行
    ws.freeze_panes = "A2"
    # 添加自动筛选
    if num_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"


def _extract_year_from_date(date_str: str) -> str:
    """从日期字符串中提取年份"""
    if not date_str:
        return "未知年份"
    # 处理多种日期格式
    date_str = str(date_str)
    for sep in ["-", "/", "."]:
        if sep in date_str:
            return date_str.split(sep)[0]
    if len(date_str) >= 4:
        return date_str[:4]
    return date_str


def write_company_excel(
    company_name: str,
    financial_data: Dict[str, List[Dict[str, Any]]],
    output_dir: str = None,
) -> str:
    """
    将单个公司的财务数据写入Excel

    参数:
        company_name: 公司名称（含股票代码）
        financial_data: 财务数据，格式:
            {"资产负债表": [...], "利润表": [...], "现金流量表": [...]}
        output_dir: 输出目录

    返回:
        生成的文件路径
    """
    if output_dir is None:
        output_dir = config.output_dir

    os.makedirs(output_dir, exist_ok=True)

    # 清理文件名中的非法字符
    safe_name = company_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    filepath = os.path.join(output_dir, f"{safe_name}_年报财务数据.xlsx")

    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 按报表类型分组数据，再将每个报表按年度分sheet
    for report_name, records in financial_data.items():
        if not records:
            # 创建空sheet
            ws = wb.create_sheet(title=f"{report_name}(无数据)")
            ws.cell(row=1, column=1, value="暂无数据")
            continue

        # 按年份分组
        year_groups: Dict[str, list] = {}
        for record in records:
            report_date = record.get("报告期", "") or record.get("报告日期", "")
            year = _extract_year_from_date(report_date)
            if year not in year_groups:
                year_groups[year] = []
            year_groups[year].append(record)

        # 为每年创建一个sheet
        for year in sorted(year_groups.keys(), reverse=True):
            year_records = year_groups[year]

            # Sheet名称限制31字符
            sheet_title = f"{report_name}_{year}年"
            if len(sheet_title) > 31:
                sheet_title = f"{year}年"

            # 如果同一年同一类型数据多行（如修正公告），只取最新一行
            if len(year_records) > 1:
                # 优先取年报（通常报告期是12-31）
                annual_records = [
                    r for r in year_records
                    if "12-31" in str(r.get("报告期", ""))
                    or "12-31" in str(r.get("报告日期", ""))
                ]
                if annual_records:
                    year_records = annual_records[:1]
                else:
                    year_records = year_records[:1]

            ws = wb.create_sheet(title=sheet_title)

            # 收集该报表所有字段（去重）
            headers = []
            seen = set()
            for record in year_records:
                for key in record:
                    if key not in seen:
                        headers.append(key)
                        seen.add(key)

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, record in enumerate(year_records, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 格式化
            _format_sheet(ws, headers, len(headers), len(year_records))

    wb.save(filepath)
    logger.info(f"Excel已保存: {filepath}")
    return filepath


def write_all_companies(
    all_data: Dict[str, Dict[str, List[Dict[str, Any]]]],
    output_dir: str = None,
    progress_callback=None,
) -> List[str]:
    """
    批量写入所有公司的Excel文件

    参数:
        all_data: 所有公司的财务数据
        output_dir: 输出目录
        progress_callback: 进度回调

    返回:
        生成的文件路径列表
    """
    if output_dir is None:
        output_dir = config.output_dir

    os.makedirs(output_dir, exist_ok=True)
    files = []

    total = len(all_data)
    for idx, (company_name, data) in enumerate(all_data.items()):
        logger.info(f"正在写入 {idx + 1}/{total}: {company_name}")
        if progress_callback:
            progress_callback(idx + 1, total, f"正在写入: {company_name}")

        filepath = write_company_excel(company_name, data, output_dir)
        files.append(filepath)

    return files