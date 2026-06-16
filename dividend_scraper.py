#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分红公告爬取模块
从东方财富 datacenter-web API 获取 A股/H股 分红送配数据
提取分配金额、股权登记日、现金红利发放日、公告日
"""

import time
import logging
from datetime import datetime
from typing import List, Dict, Optional, Callable, Tuple

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# 东方财富 datacenter-web API
EM_API_URL = "https://datacenter-web.eastmoney.com/api/data/v1/get"

# 分红送配报告名
REPORT_NAME = "RPT_SHAREBONUS_DET"

# 请求头
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://data.eastmoney.com/yjfp/",
    "Accept": "application/json, text/plain, */*",
}

# A股/H股股本缓存
_share_count_cache: Dict[str, Optional[Tuple[int, int]]] = {}
# H股派息日缓存：key = A股代码+报告年份，value = H股派息日
_h_payment_date_cache: Dict[str, str] = {}


def _get_ah_share_counts(code: str) -> Optional[Tuple[int, int]]:
    """
    获取A股/H股股本数量

    通过东方财富行情API获取A股股本，H股股本 = 总股本 - A股股本

    Returns:
        (a_shares, h_shares) 或 None（无H股）
    """
    global _share_count_cache
    if code in _share_count_cache:
        return _share_count_cache[code]

    try:
        # 先检查是否有H股
        market_prefix = "0" if code.startswith("0") or code.startswith("3") else "1"
        secid = f"{market_prefix}.{code}"

        # 获取A股行情中的股本信息
        quote_url = "https://push2.eastmoney.com/api/qt/stock/get"
        params = {
            "secid": secid,
            "fields": "f84,f85",
            "ut": "fa5fd1943c7b386f172d6893dbbd4e1f",
        }
        resp = requests.get(quote_url, params=params, headers=HEADERS, timeout=15)
        data = resp.json()
        if not data.get("data"):
            _share_count_cache[code] = None
            return None

        total_shares = int(data["data"].get("f84", 0))
        a_float = int(data["data"].get("f85", 0))

        if total_shares <= 0:
            _share_count_cache[code] = None
            return None

        # 检查是否有H股：通过F10基本资料获取H股代码
        f10_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
        f10_params = {
            "reportName": "RPT_F10_ORG_BASICINFO",
            "columns": "STR_CODEH",
            "pageNumber": 1,
            "pageSize": 1,
            "source": "WEB",
            "client": "WEB",
            "filter": f'(SECURITY_CODE="{code}")',
        }
        try:
            f10_resp = requests.get(f10_url, params=f10_params, headers=HEADERS, timeout=15)
            f10_data = f10_resp.json()
            if f10_data.get("success") and f10_data.get("result") and f10_data["result"].get("data"):
                str_codeh = f10_data["result"]["data"][0].get("STR_CODEH", "")
                if str_codeh:
                    # 有H股，计算A股/H股股本
                    # A股股本 ≈ A股流通股本（对于AH股，流通股本≈总A股）
                    # H股股本 = 总股本 - A股股本
                    a_shares = a_float
                    h_shares = total_shares - a_shares
                    if h_shares > 0:
                        result = (a_shares, h_shares)
                        _share_count_cache[code] = result
                        logger.info(f"  {code}: A股={a_shares:,}, H股={h_shares:,}")
                        return result
        except Exception:
            pass

        # 无H股
        _share_count_cache[code] = None
        return None

    except Exception as e:
        logger.debug(f"获取 {code} AH股本失败: {e}")
        _share_count_cache[code] = None
        return None


def fetch_dividend_data(
    stock_codes: List[str],
    start_year: int,
    end_year: int,
    progress_callback: Optional[Callable] = None,
) -> Tuple[List[Dict], List[str]]:
    """
    获取指定股票代码列表的分红公告数据

    Args:
        stock_codes: 股票代码列表，如 ['000001', '600036']
        start_year: 起始年份
        end_year: 结束年份
        progress_callback: 进度回调 (current, total, message)

    Returns:
        (分红记录列表, 未找到数据的股票代码列表)
    """
    all_records = []
    not_found = []
    total = len(stock_codes)

    for idx, code in enumerate(stock_codes):
        if progress_callback:
            progress_callback(idx + 1, total, f"正在查询 {code} 分红数据...")

        records = _fetch_single_stock(code, start_year, end_year)
        if records:
            all_records.extend(records)
            logger.info(f"  {code}: 获取到 {len(records)} 条分红记录")
        else:
            not_found.append(code)
            logger.warning(f"  {code}: 未获取到分红数据")

        time.sleep(0.5)  # 请求间隔

    # 为每条记录补充A股/H股分配金额
    if progress_callback:
        progress_callback(0, 0, "正在计算A股/H股分配金额...")

    for record in all_records:
        code = record["SECURITY_CODE"]
        per_share = record["CASH_DIVIDEND_PER_SHARE"]
        total_shares = record.get("TOTAL_SHARES", 0)

        ah_counts = _get_ah_share_counts(code)
        if ah_counts:
            a_shares, h_shares = ah_counts
            record["A_SHARE_AMOUNT"] = round(per_share * a_shares, 2)
            record["H_SHARE_AMOUNT"] = round(per_share * h_shares, 2)
        else:
            # 纯A股公司：A股分配金额 = 总分配金额，H股分配金额 = 0
            record["A_SHARE_AMOUNT"] = record["TOTAL_AMOUNT"]
            record["H_SHARE_AMOUNT"] = 0

    # 为有H股的公司补充H股红利派发日期
    if progress_callback:
        progress_callback(0, 0, "正在获取H股红利派发日期...")

    # 按A股代码去重，避免重复查询
    a_codes = list(set(r["SECURITY_CODE"] for r in all_records if _get_ah_share_counts(r["SECURITY_CODE"])))
    hk_dates_cache: Dict[str, Dict[str, str]] = {}

    for a_code in a_codes:
        hk_code = _lookup_hk_code(a_code)
        if hk_code:
            hk_dates_cache[a_code] = _fetch_hk_dividend_dates(hk_code)
            logger.info(f"  {a_code} → H股代码 {hk_code}，获取到 {len(hk_dates_cache[a_code])} 条派息日")
        else:
            logger.info(f"  {a_code}: 未找到H股代码")
        time.sleep(0.3)

    # 填充分红记录的H股派息日
    for record in all_records:
        a_code = record["SECURITY_CODE"]
        if a_code in hk_dates_cache:
            report_period = record.get("REPORT_PERIOD", "")
            if report_period and report_period in hk_dates_cache[a_code]:
                record["H_PAYMENT_DATE"] = hk_dates_cache[a_code][report_period]

    return all_records, not_found


def _fetch_single_stock(code: str, start_year: int, end_year: int) -> List[Dict]:
    """
    获取单只股票的分红数据

    筛选条件:
    - 方案进度为"实施分配"或"实施方案"
    - 公告日期在指定年份范围内
    """
    records = []
    page = 1

    while True:
        params = {
            "reportName": REPORT_NAME,
            "columns": "ALL",
            "pageNumber": page,
            "pageSize": 50,
            "sortColumns": "PLAN_NOTICE_DATE",
            "sortTypes": "-1",
            "source": "WEB",
            "client": "WEB",
            "filter": f'(SECURITY_CODE="{code}")',
        }

        try:
            resp = requests.get(EM_API_URL, params=params, headers=HEADERS, timeout=30)
            if resp.status_code != 200:
                break

            body = resp.json()
            if not body.get("success") or not body.get("result"):
                break

            result = body["result"]
            data = result.get("data", [])
            if not data:
                break

            for item in data:
                # 提取关键字段
                record = _parse_dividend_record(item)
                if record is None:
                    continue

                # 按公告日期筛选年份范围
                notice_date = record.get("PLAN_NOTICE_DATE", "")
                if notice_date:
                    try:
                        dt = datetime.strptime(notice_date, "%Y-%m-%d %H:%M:%S")
                        year = dt.year
                    except ValueError:
                        try:
                            dt = datetime.strptime(notice_date, "%Y-%m-%d")
                            year = dt.year
                        except ValueError:
                            year = 0
                    if year < start_year or year > end_year:
                        continue

                records.append(record)

            # 检查是否还有更多页
            total_pages = result.get("pages", 0)
            if page >= total_pages:
                break
            page += 1

        except (requests.RequestException, ValueError, KeyError) as e:
            logger.warning(f"请求 {code} 分红数据失败 (第{page}页): {e}")
            break

    return records


def _fetch_hk_dividend_dates(hk_code: str) -> Dict[str, str]:
    """
    获取H股分红派息日期

    数据源: 东方财富 PC_HKF10 分红派息API
    URL: https://emweb.securities.eastmoney.com/PC_HKF10/CorporateEvents/GetFHSPList?code=XXXXX

    返回: {报告期: 派息日, ...}  如 {"2025-12-31": "2026-07-15"}
    """
    if not hk_code:
        return {}

    cache_key = f"hk_{hk_code}"
    if cache_key in _h_payment_date_cache:
        return _h_payment_date_cache[cache_key]  # type: ignore

    result = {}
    try:
        url = "https://emweb.securities.eastmoney.com/PC_HKF10/CorporateEvents/GetFHSPList"
        params = {"code": hk_code}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Referer": f"https://emweb.securities.eastmoney.com/PC_HKF10/pages/home/index.html?code={hk_code}",
            "Accept": "application/json, text/plain, */*",
        }
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        if resp.status_code != 200:
            logger.warning(f"获取H股分红数据失败: HTTP {resp.status_code}")
            return result

        data = resp.json()
        if not data:
            return result

        # 解析分红派息列表
        records = data if isinstance(data, list) else data.get("data", [])
        for item in records:
            # 财政年度（报告期）→ 发放日（派息日）
            fiscal_year = item.get("FiscalYear", "")  # 如 "2025"
            payment_date = item.get("PaymentDate", "")  # 如 "2026-07-15"
            year_end = item.get("YearEnd", "")  # 如 "2025-12-31"
            ex_date = item.get("ExDividendDate", "")  # 除净日
            status = item.get("Status", "")

            # 只取已实施的分红
            if "实施" in str(status) or "已實施" in str(status) or status == "1":
                # 优先使用 YearEnd，其次 FiscalYear 构造
                period = year_end if year_end else f"{fiscal_year}-12-31"
                if payment_date:
                    result[period] = payment_date

        logger.info(f"  H股分红派息日期 ({hk_code}): 获取到 {len(result)} 条")
    except Exception as e:
        logger.warning(f"获取H股分红派息日期失败 ({hk_code}): {e}")

    _h_payment_date_cache[cache_key] = result  # type: ignore
    return result


def _lookup_hk_code(a_code: str) -> str:
    """
    通过A股代码查找对应的H股代码

    使用 cninfo 搜索API查询公司名称，然后匹配H股代码
    """
    if not a_code:
        return ""

    cache_key = f"hk_lookup_{a_code}"
    if cache_key in _h_payment_date_cache:
        return _h_payment_date_cache[cache_key]  # type: ignore

    result = ""
    try:
        # 先通过A股代码查询公司名称
        search_url = "http://www.cninfo.com.cn/new/information/topSearch/detailOfQuery"
        params = {"keyWord": a_code}
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "http://www.cninfo.com.cn/",
        }
        resp = requests.post(search_url, data=params, headers=headers, timeout=10)
        data = resp.json()
        items = data.get("keyBoardList") or []

        # 找到公司名称后，搜索H股代码
        company_name = ""
        for item in items:
            if item.get("code") == a_code:
                company_name = item.get("zwjc", "")
                break

        if company_name:
            # 用公司名称再次搜索，获取H股代码
            params["keyWord"] = company_name
            resp2 = requests.post(search_url, data=params, headers=headers, timeout=10)
            data2 = resp2.json()
            items2 = data2.get("keyBoardList") or []

            for item in items2:
                plate = item.get("plate", "")
                code = item.get("code", "")
                # H股市场标识：hk, HK, 港
                if any(tag in str(plate).lower() for tag in ["hk", "港", "hkg"]):
                    # 取纯数字部分
                    result = code.split(".")[0].strip()
                    if result and result.isdigit():
                        break
    except Exception as e:
        logger.debug(f"查找 {a_code} H股代码失败: {e}")

    _h_payment_date_cache[cache_key] = result  # type: ignore
    return result


def _parse_dividend_record(item: Dict) -> Optional[Dict]:
    """
    解析单条分红记录，提取关键字段

    东方财富 RPT_SHAREBONUS_DET 关键字段:
    - SECURITY_CODE: 股票代码
    - SECURITY_NAME_ABBR: 股票简称
    - PLAN_NOTICE_DATE: 预案公告日
    - NOTICE_DATE: 实施公告日
    - EQUITY_RECORD_DATE: 股权登记日
    - EX_DIVIDEND_DATE: 除权除息日（同时也是现金红利发放日）
    - PRETAX_BONUS_RMB: 每10股税前分红(元)
    - TOTAL_SHARES: 总股本
    - ASSIGN_PROGRESS: 方案进度
    - IMPL_PLAN_PROFILE: 实施方案描述
    """
    try:
        plan_status = item.get("ASSIGN_PROGRESS", "")
        # 只保留已实施的分红方案
        if "实施" not in str(plan_status):
            return None

        # 每10股税前分红
        pretax_bonus = _safe_float(item.get("PRETAX_BONUS_RMB", 0))
        # 每股分红
        per_share = pretax_bonus / 10.0 if pretax_bonus > 0 else 0

        # 总股本
        total_shares = _safe_float(item.get("TOTAL_SHARES", 0))
        # 分配金额 = 每股分红 × 总股本
        total_amount = per_share * total_shares if total_shares > 0 else 0

        return {
            "SECURITY_CODE": item.get("SECURITY_CODE", ""),
            "SECURITY_NAME": item.get("SECURITY_NAME_ABBR", ""),
            "PLAN_NOTICE_DATE": _normalize_date(item.get("PLAN_NOTICE_DATE", "")),
            "NOTICE_DATE": _normalize_date(item.get("NOTICE_DATE", "")),
            "REGIST_DATE": _normalize_date(item.get("EQUITY_RECORD_DATE", "")),
            "EX_DIVIDEND_DATE": _normalize_date(item.get("EX_DIVIDEND_DATE", "")),
            "PAYMENT_DATE": _normalize_date(item.get("EX_DIVIDEND_DATE", "")),  # 现金红利发放日≈除权除息日
            "H_PAYMENT_DATE": "",  # H股红利派发日期（东方财富API暂无此字段，需人工补充）
            "CASH_DIVIDEND_PER_SHARE": round(per_share, 4),
            "TOTAL_AMOUNT": round(total_amount, 2),
            "TOTAL_SHARES": int(total_shares),  # 总股本（用于后续计算AH拆分）
            "REPORT_TYPE": _report_type(item.get("REPORT_DATE", "")),  # 年报/中报
            "REPORT_YEAR": _report_year(item.get("REPORT_DATE", "")),  # 报告年度
            "PLAN_STATUS": str(plan_status).strip(),
            "REPORT_PERIOD": item.get("REPORT_DATE", ""),
            "IMPL_PLAN_PROFILE": item.get("IMPL_PLAN_PROFILE", ""),
        }
    except Exception as e:
        logger.debug(f"解析分红记录失败: {e}")
        return None


def _safe_float(value) -> float:
    """安全转换为浮点数"""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _normalize_date(date_str: str) -> str:
    """标准化日期格式为 YYYY-MM-DD"""
    if not date_str:
        return ""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
        try:
            dt = datetime.strptime(str(date_str).strip(), fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return str(date_str).strip()


def _report_type(report_date: str) -> str:
    """根据报告期判定报告类型"""
    if not report_date:
        return ""
    dt = _normalize_date(report_date)
    if dt.endswith("-12-31"):
        return "年报"
    elif dt.endswith("-06-30"):
        return "中报"
    elif dt.endswith("-03-31"):
        return "一季报"
    elif dt.endswith("-09-30"):
        return "三季报"
    return ""


def _report_year(report_date: str) -> str:
    """根据报告期提取报告年度"""
    if not report_date:
        return ""
    dt = _normalize_date(report_date)
    if dt:
        return dt[:4]
    return ""


def write_dividend_excel(
    records: List[Dict],
    start_year: int,
    end_year: int,
    output_dir: str = "output",
) -> str:
    """
    将分红数据写入Excel文件

    输出列:
    - 股票代码、股票名称、公告日、股权登记日、现金红利发放日、每股分红(元)、分配金额(元)
    - 按股票代码分组，每组内按公告日排序
    """
    import os

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{start_year}-{end_year}年分红公告"

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    group_font = Font(name="微软雅黑", size=10, bold=True, color="8E44AD")
    group_fill = PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid")

    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )

    # 标题行
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"{start_year}-{end_year}年 上市公司分红公告"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="8E44AD")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 表头
    headers = ["股票代码", "股票名称", "公告日", "报告年度", "报告类型", "股权登记日", "现金红利发放日", "H股红利派发日期", "每股分红(元)", "分配金额(元)", "A股分配金额(元)", "H股分配金额(元)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 按股票代码分组排序
    from collections import OrderedDict

    groups = OrderedDict()
    for rec in records:
        code = rec.get("SECURITY_CODE", "未知")
        if code not in groups:
            groups[code] = []
        groups[code].append(rec)

    current_row = 3
    for code, recs in groups.items():
        # 按公告日降序排列
        recs.sort(key=lambda x: x.get("PLAN_NOTICE_DATE", ""), reverse=True)

        name = recs[0].get("SECURITY_NAME", "") if recs else ""

        # 分组标题行
        ws.merge_cells(f"A{current_row}:L{current_row}")
        group_cell = ws.cell(row=current_row, column=1, value=f"{code}  {name}  ({len(recs)}条)")
        group_cell.font = group_font
        group_cell.fill = group_fill
        for c in range(1, 13):
            ws.cell(row=current_row, column=c).border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for rec in recs:
            ws.cell(row=current_row, column=1, value=rec.get("SECURITY_CODE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value=rec.get("SECURITY_NAME", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=2).border = thin_border

            ws.cell(row=current_row, column=3, value=rec.get("NOTICE_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=3).border = thin_border

            ws.cell(row=current_row, column=4, value=rec.get("REPORT_YEAR", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=4).border = thin_border

            ws.cell(row=current_row, column=5, value=rec.get("REPORT_TYPE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=5).border = thin_border

            ws.cell(row=current_row, column=6, value=rec.get("REGIST_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=6).border = thin_border

            ws.cell(row=current_row, column=7, value=rec.get("PAYMENT_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=7).border = thin_border

            ws.cell(row=current_row, column=8, value=rec.get("H_PAYMENT_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=8).border = thin_border

            ws.cell(row=current_row, column=9, value=rec.get("CASH_DIVIDEND_PER_SHARE", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=9).number_format = "#,##0.0000"
            ws.cell(row=current_row, column=9).border = thin_border

            ws.cell(row=current_row, column=10, value=rec.get("TOTAL_AMOUNT", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=10).number_format = "#,##0.00"
            ws.cell(row=current_row, column=10).border = thin_border

            ws.cell(row=current_row, column=11, value=rec.get("A_SHARE_AMOUNT", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=11).number_format = "#,##0.00"
            ws.cell(row=current_row, column=11).border = thin_border

            ws.cell(row=current_row, column=12, value=rec.get("H_SHARE_AMOUNT", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=12).number_format = "#,##0.00"
            ws.cell(row=current_row, column=12).border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 20

    # 冻结
    ws.freeze_panes = "A3"

    # 保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"分红公告_{start_year}-{end_year}年_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    logger.info(f"分红Excel已保存: {filepath} (共{len(records)}条记录)")

    return filepath


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    def cb(cur, total, msg):
        print(f"\r[{cur}/{total}] {msg}", end="")

    records, nf = fetch_dividend_data(["000001", "600036"], 2024, 2025, progress_callback=cb)
    print(f"\n获取 {len(records)} 条记录, 未找到: {nf}")
    if records:
        path = write_dividend_excel(records, 2024, 2025)
        print(f"已输出: {path}")