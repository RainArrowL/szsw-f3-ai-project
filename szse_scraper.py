#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深交所日度概况爬取模块
从 https://www.szse.cn/market/stock/situation/daily/index.html 获取日度概况数据
提取深市合计的成交量和成交金额
"""

import random
import time
import logging
from datetime import date, timedelta, datetime
from typing import List, Tuple, Optional

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# SZSE API 端点
SZSE_API_URL = "https://www.szse.cn/api/report/ShowReport/data"

# 请求头
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://www.szse.cn/market/stock/situation/daily/index.html",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}


def generate_trading_dates(year: int) -> List[str]:
    """
    生成指定年份的所有潜在交易日（周一至周五）
    返回格式为 YYYY-MM-DD 的日期列表
    """
    dates = []
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    current = start
    while current <= end:
        if current.weekday() < 5:  # 周一到周五
            dates.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    return dates


def fetch_daily_data(query_date: str) -> Optional[dict]:
    """
    从深交所API获取指定日期的日度概况数据

    Args:
        query_date: 查询日期，格式 YYYY-MM-DD

    Returns:
        包含日度概况数据的字典，包含 主板/创业板/深市合计 的成交量和成交金额
        如果当天不是交易日则返回 None
    """
    params = {
        "SHOWTYPE": "JSON",
        "CATALOGID": "1815",
        "txtDate": query_date,
        "random": random.random(),
    }

    try:
        resp = requests.get(SZSE_API_URL, params=params, headers=HEADERS, timeout=30)
        resp.raise_for_status()

        result = resp.json()

        # 检查返回数据结构
        if not result or not isinstance(result, list):
            return None

        first_item = result[0]
        if "error" in first_item and first_item["error"]:
            # 非交易日或API返回错误
            return None

        if "data" not in first_item or not first_item["data"]:
            return None

        # 解析数据
        data_rows = first_item["data"]
        parsed = {}

        for row in data_rows:
            # 根据常见的字段名匹配
            # 字段可能是: zqjc(证券简称), cjgs(成交量), cjje(成交金额)
            name = None
            volume = None
            amount = None

            # 尝试多种可能的字段名组合
            if "zqjc" in row:
                name = str(row["zqjc"]).replace("&nbsp;", "").strip()
            elif "name" in row:
                name = str(row["name"]).replace("&nbsp;", "").strip()

            if "cjgs" in row:
                volume = _parse_number(row["cjgs"])
            elif "volume" in row:
                volume = _parse_number(row["volume"])
            elif "VOLUME" in row:
                volume = _parse_number(row["VOLUME"])

            if "cjje" in row:
                amount = _parse_number(row["cjje"])
            elif "amount" in row:
                amount = _parse_number(row["amount"])
            elif "AMOUNT" in row:
                amount = _parse_number(row["AMOUNT"])

            if name and volume is not None and amount is not None:
                parsed[name] = {"volume": volume, "amount": amount}

        if parsed:
            return parsed

        # 如果上面的解析失败，尝试另一种数据格式
        # 某些API版本返回的是扁平结构
        return _parse_flat_format(data_rows)

    except requests.RequestException as e:
        logger.warning(f"请求 {query_date} 失败: {e}")
        return None
    except (ValueError, KeyError, IndexError) as e:
        logger.warning(f"解析 {query_date} 数据失败: {e}")
        return None


def _parse_flat_format(data_rows: list) -> Optional[dict]:
    """
    尝试解析扁平化数据格式（备用方案）
    某些API版本将所有数据放在一个扁平列表中
    """
    try:
        result = {}
        for row in data_rows:
            # 遍历所有键值对，尝试找到名称和数值
            keys = list(row.keys())
            values = list(row.values())

            name_key = None
            for k in keys:
                if k.lower() in ("zqjc", "name", "item", "type", "category"):
                    name_key = k
                    break

            if name_key:
                name = str(row[name_key]).replace("&nbsp;", "").strip()
                vol = None
                amt = None
                for k in keys:
                    if k.lower() in ("cjgs", "volume", "vol"):
                        vol = _parse_number(row[k])
                    if k.lower() in ("cjje", "amount", "amt", "turnover"):
                        amt = _parse_number(row[k])
                if name and vol is not None and amt is not None:
                    result[name] = {"volume": vol, "amount": amt}

        return result if result else None
    except Exception:
        return None


def _parse_number(value) -> Optional[float]:
    """解析数值，去除逗号和空格"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace(" ", "").strip())
    except (ValueError, TypeError):
        return None


def fetch_year_data(year: int, progress_callback=None) -> List[Tuple[str, float, float]]:
    """
    获取指定年份所有交易日的深市合计数据

    Args:
        year: 年份
        progress_callback: 进度回调函数，签名: callback(current, total, message)

    Returns:
        [(日期, 成交量(亿), 成交金额(亿元)), ...] 列表
    """
    dates = generate_trading_dates(year)
    results = []
    total = len(dates)

    for idx, d in enumerate(dates):
        if progress_callback:
            progress_callback(idx + 1, total, f"正在查询 {d}...")

        data = fetch_daily_data(d)

        if data and "深市合计" in data:
            sz_data = data["深市合计"]
            results.append((d, sz_data["volume"], sz_data["amount"]))
            logger.info(f"  {d}: 成交量={sz_data['volume']}亿, 成交金额={sz_data['amount']}亿元")
        else:
            logger.debug(f"  {d}: 非交易日或无数据")

        # 请求间隔，避免被限流
        time.sleep(0.3)

    return results


def write_szse_excel(year: int, data: List[Tuple[str, float, float]], output_dir: str = "output") -> str:
    """
    将深交所日度概况数据写入Excel文件

    输出格式：
    - 行：日期（按周分组，每周后插入小计行）
    - 列：成交量（亿）、成交金额（亿元）、日均成交金额、成交额×0.05%
    - 日均成交金额仅在每周小计行显示，计算公式：当周成交金额小计 ÷ 当周实际交易日天数
    """
    import os
    from datetime import date as date_type
    from collections import OrderedDict

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年深市日度概况"

    # ---- 样式定义 ----
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="center", vertical="center")
    date_alignment = Alignment(horizontal="center", vertical="center")

    subtotal_font = Font(name="微软雅黑", size=10, bold=True)
    subtotal_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    subtotal_alignment = Alignment(horizontal="center", vertical="center")

    summary_font = Font(name="微软雅黑", size=10, bold=True)
    summary_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )

    # ---- 标题行 ----
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"{year}年 深圳证券交易所 日度概况（深市合计）"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1A5276")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ---- 表头 ----
    headers = ["日期", "成交量（亿）", "成交金额（亿元）", "日均成交金额", "成交额×0.05%"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # ---- 按周分组 ----
    weeks = OrderedDict()
    for d_str, vol, amt in data:
        dt = date_type.fromisoformat(d_str)
        iso_year, iso_week, _ = dt.isocalendar()
        week_key = f"{iso_year}-W{iso_week:02d}"
        if week_key not in weeks:
            weeks[week_key] = []
        weeks[week_key].append((d_str, vol, amt))

    # ---- 写入数据 ----
    current_row = 3
    total_vol = 0.0
    total_amt = 0.0
    total_days = 0

    for week_key, week_data in weeks.items():
        # 本周数据
        week_vol = 0.0
        week_amt = 0.0
        week_days = len(week_data)

        for d_str, vol, amt in week_data:
            fee_005 = round(amt * 0.0005, 2)

            ws.cell(row=current_row, column=1, value=d_str).alignment = date_alignment
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value=vol).alignment = cell_alignment
            ws.cell(row=current_row, column=2).number_format = "#,##0.00"
            ws.cell(row=current_row, column=2).border = thin_border

            ws.cell(row=current_row, column=3, value=amt).alignment = cell_alignment
            ws.cell(row=current_row, column=3).number_format = "#,##0.00"
            ws.cell(row=current_row, column=3).border = thin_border

            # 日均成交金额：每日行留空
            ws.cell(row=current_row, column=4).alignment = cell_alignment
            ws.cell(row=current_row, column=4).border = thin_border

            ws.cell(row=current_row, column=5, value=fee_005).alignment = cell_alignment
            ws.cell(row=current_row, column=5).number_format = "#,##0.00"
            ws.cell(row=current_row, column=5).border = thin_border

            ws.row_dimensions[current_row].height = 20

            week_vol += vol
            week_amt += amt
            current_row += 1

        # ---- 本周小计行 ----
        avg_daily_amt = round(week_amt / week_days, 2) if week_days > 0 else 0
        fee_005_sub = round(week_amt * 0.0005, 2)

        for col in range(1, 6):
            ws.cell(row=current_row, column=col).fill = subtotal_fill
            ws.cell(row=current_row, column=col).font = subtotal_font
            ws.cell(row=current_row, column=col).alignment = subtotal_alignment
            ws.cell(row=current_row, column=col).border = thin_border

        ws.cell(row=current_row, column=1, value=f"{week_key} 小计（{week_days}个交易日）")

        ws.cell(row=current_row, column=2, value=round(week_vol, 2)).number_format = "#,##0.00"
        ws.cell(row=current_row, column=3, value=round(week_amt, 2)).number_format = "#,##0.00"
        ws.cell(row=current_row, column=4, value=avg_daily_amt).number_format = "#,##0.00"
        ws.cell(row=current_row, column=5, value=fee_005_sub).number_format = "#,##0.00"

        ws.row_dimensions[current_row].height = 22

        total_vol += week_vol
        total_amt += week_amt
        total_days += week_days
        current_row += 1

    # ---- 年度汇总行 ----
    current_row += 0  # 不额外空行，紧接最后一周小计
    year_avg_amt = round(total_amt / total_days, 2) if total_days > 0 else 0
    year_fee_005 = round(total_amt * 0.0005, 2)

    for col in range(1, 6):
        ws.cell(row=current_row, column=col).fill = summary_fill
        ws.cell(row=current_row, column=col).font = summary_font
        ws.cell(row=current_row, column=col).alignment = subtotal_alignment
        ws.cell(row=current_row, column=col).border = thin_border

    ws.cell(row=current_row, column=1, value=f"{year}年 合计（{total_days}个交易日）")
    ws.cell(row=current_row, column=2, value=round(total_vol, 2)).number_format = "#,##0.00"
    ws.cell(row=current_row, column=3, value=round(total_amt, 2)).number_format = "#,##0.00"
    ws.cell(row=current_row, column=4, value=year_avg_amt).number_format = "#,##0.00"
    ws.cell(row=current_row, column=5, value=year_fee_005).number_format = "#,##0.00"
    ws.row_dimensions[current_row].height = 24

    # ---- 列宽 ----
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # ---- 冻结窗格（冻结标题+表头） ----
    ws.freeze_panes = "A3"

    # 保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"深交所日度概况_{year}年_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    logger.info(f"Excel已保存: {filepath}")

    return filepath


if __name__ == "__main__":
    # 测试
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    def progress_cb(cur, total, msg):
        print(f"\r[{cur}/{total}] {msg}", end="")

    data = fetch_year_data(2025, progress_callback=progress_cb)
    print(f"\n共获取 {len(data)} 个交易日数据")
    if data:
        path = write_szse_excel(2025, data)
        print(f"已输出: {path}")