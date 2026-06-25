#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公募基金管理人名录爬取模块
数据来源：中国证券投资基金业协会 (AMAC)
API: https://www.amac.org.cn/portal/front/mutualFund/findMutualFundHousePage
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional

import requests

logger = logging.getLogger(__name__)

AMAC_API_URL = "https://www.amac.org.cn/portal/front/mutualFund/findMutualFundHousePage"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://www.amac.org.cn/fwdt/wyc/jgcprycx/jgcx/gmjjglrml/",
    "Accept": "application/json, text/plain, */*",
}

# 字段映射：英文 → 中文
FIELD_CN = {
    "lineId": "序号",
    "houseName": "公司名称",
    "registerAddr": "注册地",
    "officeAddr": "辖区",
    "website": "官方网址",
    "phone": "客服电话",
}


def fetch_fund_manager_list() -> Optional[List[Dict[str, str]]]:
    """
    从 AMAC 获取公募基金管理人名录全量数据

    返回：
        字典列表，字段已翻译为中文
    """
    try:
        resp = requests.get(
            AMAC_API_URL,
            params={"pageNo": 1, "pageSize": 500},
            headers=HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        body = resp.json()

        if body.get("code") != 200:
            logger.error(f"AMAC API 返回错误: {body}")
            return None

        data = body.get("data", {})
        if data.get("errcode") != 0:
            logger.error(f"AMAC API 业务错误: {data.get('msg')}")
            return None

        inner = data.get("data", {})
        data_list = inner.get("dataList", [])
        total = inner.get("total", 0)

        logger.info(f"获取到 {len(data_list)} 条公募基金管理人记录 (共 {total} 条)")

        # 翻译字段为中文
        result = []
        for item in data_list:
            row = {}
            for en_key, cn_key in FIELD_CN.items():
                val = item.get(en_key, "")
                # 清理换行符
                if isinstance(val, str):
                    val = val.replace("\n", " / ").replace("\r", "")
                row[cn_key] = val
            result.append(row)

        return result

    except requests.RequestException as e:
        logger.error(f"请求 AMAC API 失败: {e}")
        return None
    except Exception as e:
        logger.error(f"解析 AMAC 数据失败: {e}")
        return None


def write_amac_excel(records: List[Dict[str, str]], output_dir: str) -> str:
    """
    将公募基金管理人名录写入 Excel

    参数：
        records: 翻译后的记录列表
        output_dir: 输出目录

    返回：
        文件路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    filepath = str(Path(output_dir) / "公募基金管理人名录.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "公募基金管理人名录"

    # 表头
    headers = list(FIELD_CN.values())
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B5151D", end_color="B5151D", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, record in enumerate(records, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # 列宽
    col_widths = [6, 30, 10, 10, 28, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"

    wb.save(filepath)
    logger.info(f"公募基金管理人名录已保存: {filepath} (共 {len(records)} 条)")
    return filepath