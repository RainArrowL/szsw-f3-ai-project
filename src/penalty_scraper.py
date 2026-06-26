"""
金融机构行政处罚信息爬取模块
数据来源：
  1. 国家金融监督管理总局（NFRA）— 总局机关 + 派出机构
  2. 中国人民银行（PBC）— 行政处罚公示
  3. 中国证监会（CSRC）— 行政处罚决定
"""
import re
import logging
import requests
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
from urllib.parse import urljoin, urlparse

logger = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9",
}

# ═══════════════════════════════════════════════════════════════
# NFRA（国家金融监督管理总局）行政处罚
# ═══════════════════════════════════════════════════════════════

# 总局机关行政处罚列表
NFRA_ZONGJU_LIST = (
    "https://www.nfra.gov.cn/cn/view/pages/ItemList.html"
    "?itemPId=923&itemId=4113&itemUrl=ItemListRightList.html"
    "&itemName=%E6%80%BB%E5%B1%80%E6%9C%BA%E5%85%B3"
)
# 派出机构行政处罚列表
NFRA_PAICHU_LIST = (
    "https://www.nfra.gov.cn/cn/view/pages/ItemList.html"
    "?itemPId=923&itemId=4293&itemUrl=ItemListRightList.html"
    "&itemName=%E6%B4%BE%E5%87%BA%E6%9C%BA%E6%9E%84"
)
NFRA_DETAIL_BASE = "https://www.nfra.gov.cn/cn/view/pages/ItemDetail.html"
NFRA_API_BASE = "https://www.nfra.gov.cn/cbircweb"


def _fetch_html(url: str, timeout: int = 30) -> Optional[str]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.encoding = "utf-8"
        if resp.status_code == 200:
            return resp.text
        logger.warning(f"请求失败 {url}: HTTP {resp.status_code}")
    except Exception as e:
        logger.warning(f"请求异常 {url}: {e}")
    return None


def _fetch_json(url: str, params: dict = None, timeout: int = 30) -> Optional[dict]:
    try:
        resp = requests.get(url, params=params, headers=HEADERS, timeout=timeout)
        if resp.status_code == 200:
            return resp.json()
        logger.warning(f"API请求失败 {url}: HTTP {resp.status_code}")
    except Exception as e:
        logger.warning(f"API请求异常 {url}: {e}")
    return None


def _parse_html_table(html: str) -> List[List[str]]:
    """从HTML解析表格，返回二维数组（跳过表头）"""
    tr_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL)
    td_pattern = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.DOTALL)
    tag_pattern = re.compile(r"<[^>]+>")
    ws_pattern = re.compile(r"\s+")

    rows = []
    trs = tr_pattern.findall(html)
    for tr in trs:
        cells = td_pattern.findall(tr)
        if not cells:
            continue
        row = [ws_pattern.sub(" ", tag_pattern.sub("", c)).strip() for c in cells]
        if row and any(c for c in row):
            rows.append(row)
    return rows


def _extract_nfra_docids(html: str) -> List[str]:
    """从NFRA列表页提取所有docId"""
    docids = []
    pattern = re.compile(r'docId=(\d+)')
    seen = set()
    for m in pattern.finditer(html):
        did = m.group(1)
        if did not in seen:
            seen.add(did)
            docids.append(did)
    return docids


def _extract_nfra_docids(html: str) -> List[str]:
    """从NFRA列表页提取所有docId使用API"""
    docids = []
    pattern = re.compile(r'docId=(\d+)')
    seen = set()
    for m in pattern.finditer(html):
        did = m.group(1)
        if did not in seen:
            seen.add(did)
            docids.append(did)
    return docids


def _parse_nfra_detail_from_api(doc_id: str) -> List[Dict]:
    """直接通过API获取NFRA行政处罚详情"""
    results = []
    api_url = f"{NFRA_API_BASE}/DocInfo/SelectByDocId"
    api_data = _fetch_json(api_url, {"docId": doc_id})
    if not api_data or api_data.get("rptCode") != 200:
        return []

    doc_data = api_data.get("data", {})
    html_content = doc_data.get("docClob", "")
    if not html_content:
        return []

    tables = _parse_html_table(html_content)
    for row in tables:
        if not row:
            continue
        first_cell = row[0] if row else ""
        if "序号" in first_cell or "当事人" in first_cell or "作出决定" in first_cell:
            continue
        if not first_cell or not first_cell.strip().isdigit():
            continue

        result = {
            "序号": row[0] if len(row) > 0 else "",
            "当事人名称": row[1] if len(row) > 1 else "",
            "主要违法违规行为": row[2] if len(row) > 2 else "",
            "行政处罚内容": row[3] if len(row) > 3 else "",
            "作出决定机关": row[4] if len(row) > 4 else "",
        }
        results.append(result)

    pub_date = doc_data.get("builddate", "")
    if results and pub_date:
        for r in results:
            r["发布日期"] = pub_date
    return results


def _fetch_nfra_docids(item_id: str, max_pages: int = 5) -> List[Dict]:
    """通过API获取NFRA文档列表"""
    docids = []
    api_url = f"{NFRA_API_BASE}/DocInfo/SelectDocByItemIdAndChild"
    page_size = 18

    for page_idx in range(1, max_pages + 1):
        api_data = _fetch_json(api_url, {
            "itemId": item_id,
            "pageSize": page_size,
            "pageIndex": page_idx,
        })
        if not api_data or api_data.get("rptCode") != 200:
            break
        rows = api_data.get("data", {}).get("rows", [])
        if not rows:
            break
        for row in rows:
            docids.append({
                "docId": row.get("docId"),
                "docTitle": row.get("docTitle", ""),
                "builddate": row.get("builddate", ""),
                "publishDate": row.get("publishDate", ""),
            })
    return docids


def fetch_nfra_penalty(max_pages: int = 5) -> List[Dict]:
    """获取NFRA行政处罚信息（总局机关 + 派出机构）"""
    all_penalties = []

    for source_name, item_id in [
        ("总局机关", "4113"),
        ("派出机构", "4293"),
    ]:
        logger.info(f"正在获取NFRA {source_name} 行政处罚列表...")
        doc_list = _fetch_nfra_docids(item_id, max_pages)
        logger.info(f"NFRA {source_name}: 找到 {len(doc_list)} 条记录")

        for doc in doc_list:
            did = doc["docId"]
            records = _parse_nfra_detail_from_api(did)
            for r in records:
                r["来源"] = f"国家金融监督管理总局{source_name}"
                r["文号"] = ""
                r["处罚金额"] = _extract_penalty_amount(r.get("行政处罚内容", ""))
            all_penalties.extend(records)

    logger.info(f"NFRA 行政处罚总计: {len(all_penalties)} 条")
    return all_penalties


# ═══════════════════════════════════════════════════════════════
# PBC（中国人民银行）行政处罚
# ═══════════════════════════════════════════════════════════════

PBC_LIST_URL = (
    "https://www.pbc.gov.cn/zhengwugongkai/4081330/4081344/"
    "4081407/4081705/index.html"
)
PBC_BASE = "https://www.pbc.gov.cn"


def _extract_pbc_article_ids(html: str) -> List[str]:
    """从PBC列表页提取所有文章ID"""
    pattern = re.compile(
        r'/zhengwugongkai/4081330/4081344/4081407/4081705/(\d+/index\.html)'
    )
    ids = []
    seen = set()
    for m in pattern.finditer(html):
        aid = m.group(1)
        if aid not in seen:
            seen.add(aid)
            ids.append(aid)
    return ids


def _parse_pbc_detail(html: str) -> List[Dict]:
    """解析PBC行政处罚详情页，返回多条记录"""
    results = []

    # 提取发布日期
    date_match = re.search(r'发布时间[：:]\s*(\d{4})年(\d{1,2})月(\d{1,2})日', html)
    if date_match:
        pub_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
    else:
        pub_date = ""

    # 提取文号（标题）
    title_match = re.search(r'<title>(.*?)</title>', html, re.DOTALL)
    wenhao = ""
    if title_match:
        wenhao = title_match.group(1).strip()

    # 解析HTML表格
    tables = _parse_html_table(html)
    for row in tables:
        if not row:
            continue
        first_cell = row[0] if row else ""
        # 跳过表头
        if "序号" in first_cell or "当事人名称" in first_cell:
            continue
        if not first_cell or not first_cell.strip().isdigit():
            continue

        result = {
            "当事人名称": row[1] if len(row) > 1 else "",
            "文号": row[2] if len(row) > 2 else wenhao,
            "主要违法违规行为": row[3] if len(row) > 3 else "",
            "行政处罚内容": row[4] if len(row) > 4 else "",
            "作出决定机关": row[5] if len(row) > 5 else "中国人民银行",
            "发布日期": row[6] if len(row) > 6 else pub_date,
            "来源": "中国人民银行",
            "处罚金额": "",
        }
        result["处罚金额"] = _extract_penalty_amount(result.get("行政处罚内容", ""))
        results.append(result)

    return results


def fetch_pbc_penalty(max_pages: int = 10) -> List[Dict]:
    """获取PBC（央行）行政处罚信息"""
    logger.info("正在获取中国人民银行行政处罚列表...")
    html = _fetch_html(PBC_LIST_URL)
    if not html:
        return []

    article_ids = _extract_pbc_article_ids(html)
    logger.info(f"PBC 行政处罚: 找到 {len(article_ids)} 条记录")

    results = []
    for aid in article_ids[:max_pages]:
        url = f"{PBC_BASE}/zhengwugongkai/4081330/4081344/4081407/4081705/{aid}"
        detail_html = _fetch_html(url)
        if not detail_html:
            continue
        records = _parse_pbc_detail(detail_html)
        results.extend(records)

    logger.info(f"PBC 行政处罚总计: {len(results)} 条")
    return results


# ═══════════════════════════════════════════════════════════════
# CSRC（证监会）行政处罚
# ═══════════════════════════════════════════════════════════════

CSRC_PENALTY_LIST = (
    "http://www.csrc.gov.cn/csrc/index.shtml"
)

# CSRC 行政处罚决定书最新已知链接（index.shtml 为 JS 渲染，requests 无法获取列表）
# 通过 WebSearch 定期更新
CSRC_KNOWN_PENALTY_URLS = [
    "http://www.csrc.gov.cn/csrc/c101928/c7637445/content.shtml",
    "http://www.csrc.gov.cn/csrc/c101928/c7637442/content.shtml",
    "http://www.csrc.gov.cn/csrc/c101928/c7637439/content.shtml",
    "http://www.csrc.gov.cn/csrc/c101928/c7634421/content.shtml",
    "http://www.csrc.gov.cn/csrc/c101928/c7631484/content.shtml",
    "http://www.csrc.gov.cn/csrc/c101928/c7630121/content.shtml",
]


def _extract_csrc_penalty_links(html: str) -> List[Dict]:
    """从CSRC行政处罚列表页提取链接和日期"""
    results = []
    # 匹配行政处罚决定书链接：内容区域中包含 xingzhengchufa 或 c101928 的链接
    pattern = re.compile(
        r'<a[^>]*href="([^"]*?(?:content|detail)[^"]*\.shtml)"[^>]*>\s*(.*?)\s*</a>',
        re.DOTALL
    )
    date_pattern = re.compile(r'(\d{4}-\d{2}-\d{2})')

    # 只在主要内容区域查找
    main_match = re.search(
        r'<div[^>]*class="[^"]*list[^"]*"[^>]*>(.*?)</div>',
        html, re.DOTALL
    )
    search_html = main_match.group(1) if main_match else html

    for m in pattern.finditer(search_html):
        href = m.group(1)
        title = re.sub(r'<[^>]+>', '', m.group(2)).strip()

        # 过滤非行政处罚链接
        if not title or len(title) < 5:
            continue
        if any(kw in title for kw in ["联系", "关于", "指南", "导航", "首页", "更多", "地图"]):
            continue

        if not href.startswith('http'):
            href = urljoin("http://www.csrc.gov.cn", href)

        # 在同级块中查找日期
        pos = m.start()
        nearby = search_html[max(0, pos - 300):pos + 500]
        date_match = date_pattern.search(nearby)
        pub_date = date_match.group(1) if date_match else ""

        results.append({"url": href, "title": title, "date": pub_date})
    return results


def _parse_csrc_penalty_detail(html: str) -> Optional[Dict]:
    """解析CSRC行政处罚决定书详情"""
    result = {
        "当事人名称": "",
        "文号": "",
        "行政处罚内容": "",
        "作出决定机关": "中国证监会",
        "发布日期": "",
        "来源": "中国证监会",
        "处罚金额": "",
    }

    # 提取标题
    title_match = re.search(r'<title>(.*?)</title>', html, re.DOTALL)
    if title_match:
        title = title_match.group(1).strip()
        # 去除尾部 "_中国证券监督管理委员会"
        title = re.sub(r'_中国证券监督管理委员会$', '', title).strip()
        result["当事人名称"] = title

    # 提取文号（从detail-news区域）
    content_match = re.search(
        r'<div[^>]*class="[^"]*detail-news[^"]*"[^>]*>(.*?)</div>',
        html, re.DOTALL
    )
    if content_match:
        content_html = content_match.group(1)
        # 提取文号
        wh_match = re.search(r'〔(\d{4})〕(\d+)号', content_html)
        if wh_match:
            result["文号"] = f"〔{wh_match.group(1)}〕{wh_match.group(2)}号"
        # 提取完整文本
        text = re.sub(r'<[^>]+>', '', content_html)
        text = re.sub(r'\s+', ' ', text).strip()
        result["行政处罚内容"] = text[:500]

    # 提取发布日期
    date_match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', html)
    if date_match:
        result["发布日期"] = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"

    result["处罚金额"] = _extract_penalty_amount(result["行政处罚内容"])
    return result


def fetch_csrc_penalty(max_pages: int = 10) -> List[Dict]:
    """获取CSRC行政处罚信息"""
    logger.info("正在获取中国证监会行政处罚列表...")

    results = []
    for url in CSRC_KNOWN_PENALTY_URLS[:max_pages]:
        detail_html = _fetch_html(url)
        if not detail_html:
            continue
        record = _parse_csrc_penalty_detail(detail_html)
        if record and record["当事人名称"]:
            results.append(record)

    logger.info(f"CSRC 行政处罚总计: {len(results)} 条")
    return results


# ═══════════════════════════════════════════════════════════════
# 通用工具函数
# ═══════════════════════════════════════════════════════════════

def _extract_penalty_amount(text: str) -> str:
    """从文本中提取罚款金额"""
    patterns = [
        r'罚款\s*([\d,.]+)\s*万?元',
        r'罚[没款]\s*([\d,.]+)\s*万?元',
        r'[没罚]\s*([\d,.]+)\s*万?元',
    ]
    for p in patterns:
        match = re.search(p, text)
        if match:
            amount = match.group(1).replace(',', '')
            return f"{amount}万元" if '万' not in match.group(0) else f"{amount}万元"
    return ""


def fetch_all_penalty(max_per_source: int = 10) -> Dict[str, List[Dict]]:
    """获取所有来源的行政处罚信息"""
    return {
        "nfra": fetch_nfra_penalty(max_per_source),
        "pbc": fetch_pbc_penalty(max_per_source),
        "csrc": fetch_csrc_penalty(max_per_source),
    }


def write_penalty_excel(
    data: Dict[str, List[Dict]],
    output_dir: str = "output",
) -> str:
    """将行政处罚信息写入Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cf = Font(name="微软雅黑", size=10)
    align = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["序号", "来源", "当事人名称", "文号", "行政处罚内容", "处罚金额", "发布日期"]

    first = True
    for key, label in [
        ("nfra", "金监总局处罚"),
        ("pbc", "央行处罚"),
        ("csrc", "证监会处罚"),
    ]:
        items = data.get(key, [])
        if first:
            ws = wb.active
            ws.title = label
            first = False
        else:
            ws = wb.create_sheet(title=label)

        # 写表头
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf
            c.fill = hfill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        # 写数据
        for i, item in enumerate(items, 1):
            vals = [
                i,
                item.get("来源", ""),
                item.get("当事人名称", ""),
                item.get("文号", ""),
                item.get("行政处罚内容", ""),
                item.get("处罚金额", ""),
                item.get("发布日期", ""),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=i + 1, column=col, value=val)
                c.font = cf
                c.alignment = align
                c.border = border

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 14

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"金融机构行政处罚信息_{timestamp}.xlsx"
    filepath = str(Path(output_dir) / filename)
    wb.save(filepath)
    logger.info(f"行政处罚Excel已保存: {filepath}")
    return filepath