"""
金融机构法人名录爬取模块
获取国家金融监督管理总局的银行保险法人名单、证监会的证券基金公司名单
"""
import re
import logging
import requests
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin

logger = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9",
}

# ── NFRA 银行保险法人名单 ───────────────────────────────────
# NFRA 网站是 AngularJS SPA，通过内部 API 链动态发现 PDF 链接。
# API 导航路径：政务信息(itemId=923) → 政府信息公开(924) → 机构监管(862) → 综合(863)
# 以下为静态备用 URL（当动态发现失败时使用）
NFRA_BANK_PDF_URL_FALLBACK = (
    "https://www.nfra.gov.cn/chinese/docfile/2025/"
    "86c58b1ad810422c8fa6c6d0107f1626.pdf"
)
NFRA_INSURANCE_PDF_URL_FALLBACK = (
    "https://www.nfra.gov.cn/chinese/docfile/2025/"
    "2a78efc6d162484f8dfb8d0388b00320.pdf"
)

NFRA_BASE_URL = "https://www.nfra.gov.cn"

# ── CSRC 证券基金期货公司名录 ───────────────────────────────────
# 上海辖区证券公司名录（上海局汇总全国证券公司）
CSRC_SECURITIES_URL = (
    "http://www.csrc.gov.cn/shanghai/c103854/c7637721/content.shtml"
)
# 上海辖区基金管理公司名录
CSRC_FUND_URL = (
    "http://www.csrc.gov.cn/shanghai/c103856/c7639412/content.shtml"
)
# 证监会公开的期货公司名录（最新）
CSRC_FUTURES_URL = (
    "http://www.csrc.gov.cn/csrc/c101920/c1039268/content.shtml"
)


def _fetch_html(url: str, timeout: int = 30) -> Optional[str]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.encoding = "utf-8"
        if resp.status_code == 200:
            return resp.text
        logger.warning(f"请求失败 {url}: HTTP {resp.status_code}")
    except Exception as e:
        logger.warning(f"请求异常 {url}: {e}")
    return None


def _fetch_json(url: str, timeout: int = 30) -> Optional[dict]:
    """请求 JSON API，返回解析后的 dict"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.encoding = "utf-8"
        if resp.status_code == 200:
            return resp.json()
        logger.warning(f"JSON API 请求失败 {url}: HTTP {resp.status_code}")
    except Exception as e:
        logger.warning(f"JSON API 请求异常 {url}: {e}")
    return None


def _download_file(url: str, timeout: int = 60) -> Optional[bytes]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code == 200:
            return resp.content
        logger.warning(f"下载失败 {url}: HTTP {resp.status_code}")
    except Exception as e:
        logger.warning(f"下载异常 {url}: {e}")
    return None


def _parse_html_table(html: str) -> List[List[str]]:
    """从HTML解析表格，返回二维数组"""
    rows = []
    tr_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL)
    td_pattern = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.DOTALL)
    tag_pattern = re.compile(r"<[^>]+>")
    ws_pattern = re.compile(r"\s+")

    trs = tr_pattern.findall(html)
    for tr in trs:
        cells = td_pattern.findall(tr)
        if not cells:
            continue
        row = [ws_pattern.sub(" ", tag_pattern.sub("", c)).strip() for c in cells]
        if row and any(c for c in row):
            rows.append(row)
    return rows


def _find_xlsx_url(html: str, base_url: str) -> Optional[str]:
    """从HTML页面中查找附件xlsx/xls链接"""
    for pattern in [r'href="([^"]+\.xlsx)"', r'href="([^"]+\.xls)"']:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            return urljoin(base_url, match.group(1))
    return None


def _find_pdf_url(html: str, base_url: str) -> Optional[str]:
    """从HTML页面中查找附件PDF链接"""
    pattern = re.compile(r'href="([^"]+\.pdf)"', re.IGNORECASE)
    for match in pattern.finditer(html):
        return urljoin(base_url, match.group(1))
    return None


def _parse_pdf_table(pdf_bytes: bytes) -> List[List[str]]:
    """从PDF二进制数据中提取表格"""
    from io import BytesIO

    # 尝试 pdfplumber
    try:
        import pdfplumber

        rows = []
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row and any(c for c in row if c):
                            rows.append([str(c).strip() if c else "" for c in row])
        return rows
    except ImportError:
        logger.warning("pdfplumber未安装，尝试PyPDF2")

    # 降级：PyPDF2
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(BytesIO(pdf_bytes))
        text = ""
        for page in reader.pages:
            extracted = page.extract_text()
            if extracted:
                text += extracted + "\n"

        lines = text.strip().split("\n")
        rows = []
        for line in lines:
            parts = re.split(r"\s{2,}", line.strip())
            if parts and any(p for p in parts):
                rows.append(parts)
        return rows
    except ImportError:
        logger.warning("PyPDF2也未安装，无法解析PDF")
        return []


def _parse_xlsx(data: bytes) -> List[List[str]]:
    """解析Excel二进制数据（支持xlsx和xls格式）"""
    from io import BytesIO

    # 尝试 openpyxl (xlsx)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return rows
    except Exception:
        pass

    # 尝试 xlrd (xls)
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            rows.append([str(ws.cell_value(r, c)) if ws.cell_value(r, c) != "" else ""
                         for c in range(ws.ncols)])
        return rows
    except ImportError:
        logger.warning("xlrd未安装，无法解析.xls文件")
    except Exception as e:
        logger.warning(f"解析Excel失败: {e}")

    return []


def _is_header_row(row: List[str], header_keywords: List[str] = None,
                   name_col_index: int = 1) -> bool:
    """判断是否为表头行或无效行（需要跳过）

    name_col_index: 名称列在行中的索引（默认第1列即索引1）
    """
    if header_keywords is None:
        header_keywords = ["序号", "中文全称", "机构名称", "公司名称", "单位名称",
                           "名称", "序号", "英文全称"]
    if not row:
        return True
    name_col = row[name_col_index] if len(row) > name_col_index else row[0] if row else ""
    # 空名称
    if not name_col or not name_col.strip():
        return True
    # 表头关键词
    for kw in header_keywords:
        if kw in name_col:
            return True
    # 注释行（如"本月无变化"、"注："等）
    if re.match(r'^(本月|注[：:]|说明|备注|截止)', name_col):
        return True
    # 非中文名称（纯英文/数字/特殊字符）
    if not re.search(r'[\u4e00-\u9fff]', name_col):
        return True
    return False


def _chunk_tables(tables: List[List[str]]) -> List[List[List[str]]]:
    """将表格拆分为多个逻辑块"""
    if not tables:
        return []
    chunks = []
    current = []
    header_keywords = ["序号", "中文全称", "机构名称", "公司名称", "名称"]
    for row in tables:
        row_str = "".join(row)
        if any(kw in row_str for kw in header_keywords) and current:
            if len(current) > 1:
                chunks.append(current)
            current = [row]
        else:
            current.append(row)
    if current:
        chunks.append(current)
    return chunks


def _discover_nfra_pdf_urls() -> List[Tuple[str, str]]:
    """通过 NFRA 内部 API 链动态发现最新的银行/保险法人名单 PDF 链接

    API 导航路径（通过 itemId 链）：
      www.nfra.gov.cn → 政务信息(923) → 政府信息公开(924) → 机构监管(862) → 综合(863)

    综合栏目下的文档列表中，通过标题匹配找到：
      - "银行业金融机构法人名单"（排除"外国银行"等非主线名单）
      - "保险机构法人名单"（排除"保险中介"、"保险专业中介"等）

    Returns:
        [(label, pdf_url), ...] 列表，如 [("银行业金融机构法人名单", "https://..."), ...]
    """
    logger.info("正在动态发现 NFRA 银行保险法人名单 PDF 链接...")

    # 综合栏目的 itemId 已知且稳定（由 API 导航路径确定）
    item_id = "863"

    # 获取「综合」栏目下的文档列表
    list_url = (
        f"{NFRA_BASE_URL}/cbircweb/DocInfo/SelectDocByItemIdAndChild"
        f"?itemId={item_id}&pageSize=50&pageIndex=1"
    )
    data = _fetch_json(list_url)
    if not data or "data" not in data:
        logger.warning("无法获取综合栏目文档列表，回退到静态 URL")
        return []

    rows = data["data"].get("rows", [])
    logger.info(f"综合栏目共 {len(rows)} 条文档")

    # 按标题匹配银行和保险名单
    bank_doc_id = None
    insurance_doc_id = None

    for row in rows:
        title = row.get("docTitle", "")
        doc_id = str(row.get("docId", ""))

        if not bank_doc_id:
            if "银行业金融机构法人名单" in title and "外国银行" not in title:
                bank_doc_id = doc_id
                logger.info(f"找到银行名单: {title} (docId={doc_id})")

        if not insurance_doc_id:
            if "保险机构法人名单" in title and "中介" not in title:
                insurance_doc_id = doc_id
                logger.info(f"找到保险名单: {title} (docId={doc_id})")

        if bank_doc_id and insurance_doc_id:
            break

    if not bank_doc_id and not insurance_doc_id:
        logger.warning("未在综合栏目中找到银行/保险法人名单")
        return []

    # 获取文档详情，提取附件 CDN PDF URL
    results = []
    for label, doc_id in [("银行业金融机构法人名单", bank_doc_id),
                           ("保险机构法人名单", insurance_doc_id)]:
        if not doc_id:
            logger.warning(f"未找到 {label} 的 docId")
            continue

        detail_url = f"{NFRA_BASE_URL}/cbircweb/DocInfo/SelectByDocId?docId={doc_id}"
        detail = _fetch_json(detail_url)
        if not detail or "data" not in detail:
            logger.warning(f"获取 {label} 详情失败")
            continue

        doc_data = detail["data"]
        attachments = doc_data.get("attachmentInfoVOList", [])
        if not attachments:
            logger.warning(f"{label} 没有附件")
            continue

        # 附件 PDF URL 在 urlOtherName 字段
        pdf_rel_path = attachments[0].get("urlOtherName", "")
        if not pdf_rel_path:
            logger.warning(f"{label} 附件 URL 为空")
            continue

        pdf_url = urljoin(NFRA_BASE_URL, pdf_rel_path)
        logger.info(f"{label} PDF URL: {pdf_url}")
        results.append((label, pdf_url))

    return results


def download_nfra_pdfs(output_dir: str = "output") -> List[str]:
    """动态发现并下载 NFRA 银行保险法人名单 PDF 文件（不转 Excel）

    通过 NFRA 内部 API 链动态发现最新 PDF 链接：
      www.nfra.gov.cn → 政务信息 → 法定主动公开内容 → 机构监管 → 综合
      → 找到「银行业金融机构法人名单」和「保险机构法人名单」PDF

    Returns:
        下载的 PDF 文件路径列表
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # 动态发现 PDF 链接
    discovered = _discover_nfra_pdf_urls()

    # 如果动态发现失败，回退到静态备用 URL
    if not discovered:
        logger.warning("动态发现失败，使用静态备用 URL")
        discovered = [
            ("银行业金融机构法人名单", NFRA_BANK_PDF_URL_FALLBACK),
            ("保险机构法人名单", NFRA_INSURANCE_PDF_URL_FALLBACK),
        ]

    files = []
    for label, url in discovered:
        logger.info(f"正在下载 {label} PDF: {url}")
        pdf_bytes = _download_file(url, timeout=120)
        if not pdf_bytes:
            logger.warning(f"下载失败: {label}")
            continue

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{label}_{timestamp}.pdf"
        filepath = Path(output_dir) / filename
        filepath.write_bytes(pdf_bytes)
        logger.info(f"{label} PDF 已保存: {filepath} ({len(pdf_bytes)} bytes)")
        files.append(str(filepath))

    return files


def download_csrc_to_combined_xlsx(output_dir: str = "output") -> str:
    """下载 CSRC 证券/基金/期货公司名录附件，合并为一个 xlsx 文件

    从 CSRC 页面找到附件链接（xlsx/xls），下载原始文件后合并到一个 xlsx：
      - Sheet「证券公司名录」— 保留原始格式
      - Sheet「基金管理公司名录」— 保留原始格式
      - Sheet「期货公司名录」— 保留原始格式

    Returns:
        合并后的 xlsx 文件路径，失败返回空字符串
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    sources = [
        ("证券公司名录", CSRC_SECURITIES_URL),
        ("基金管理公司名录", CSRC_FUND_URL),
        ("期货公司名录", CSRC_FUTURES_URL),
    ]

    wb_out = Workbook()
    # 删除默认的 Sheet
    wb_out.remove(wb_out.active)

    for label, page_url in sources:
        logger.info(f"正在获取 {label} 页面: {page_url}")
        html = _fetch_html(page_url)
        if not html:
            logger.warning(f"无法访问 {label} 页面")
            continue

        attachment_url = _find_xlsx_url(html, page_url)
        if not attachment_url:
            logger.warning(f"{label} 页面未找到附件链接")
            continue

        logger.info(f"正在下载 {label}: {attachment_url}")
        data = _download_file(attachment_url, timeout=120)
        if not data:
            logger.warning(f"下载 {label} 失败")
            continue

        ext = Path(attachment_url.split("?")[0]).suffix.lower()
        if ext == ".xlsx":
            _copy_xlsx_sheet(wb_out, data, label)
        else:
            _copy_xls_sheet(wb_out, data, label)

    if not wb_out.sheetnames:
        logger.warning("没有成功下载任何名录")
        return ""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"证券基金期货公司名录_{timestamp}.xlsx"
    filepath = Path(output_dir) / filename
    wb_out.save(str(filepath))
    logger.info(f"合并名录已保存: {filepath}")
    return str(filepath)


def _copy_xlsx_sheet(wb_out, data: bytes, sheet_name: str):
    """将 xlsx 文件的第一个 sheet 复制到目标 workbook，保留格式"""
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from copy import copy

    wb_src = load_workbook(BytesIO(data))
    ws_src = wb_src.active
    ws_out = wb_out.create_sheet(title=sheet_name)

    # 复制合并单元格
    for merged_range in ws_src.merged_cells.ranges:
        ws_out.merge_cells(str(merged_range))

    # 复制行高和列宽
    for row_idx in range(1, ws_src.max_row + 1):
        if ws_src.row_dimensions[row_idx].height:
            ws_out.row_dimensions[row_idx].height = ws_src.row_dimensions[row_idx].height
    for col_idx in range(1, ws_src.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if ws_src.column_dimensions[col_letter].width:
            ws_out.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width

    # 复制每个单元格的值和样式
    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    wb_src.close()
    logger.info(f"已复制 {sheet_name}（xlsx，{ws_src.max_row} 行 × {ws_src.max_column} 列）")


def _copy_xls_sheet(wb_out, data: bytes, sheet_name: str):
    """将 xls 文件的第一个 sheet 数据写入目标 workbook"""
    from io import BytesIO
    import xlrd

    wb_src = xlrd.open_workbook(file_contents=data)
    ws_src = wb_src.sheet_by_index(0)
    ws_out = wb_out.create_sheet(title=sheet_name)

    for r in range(ws_src.nrows):
        for c in range(ws_src.ncols):
            cell_value = ws_src.cell_value(r, c)
            ws_out.cell(row=r + 1, column=c + 1, value=cell_value if cell_value != "" else None)

    logger.info(f"已复制 {sheet_name}（xls，{ws_src.nrows} 行 × {ws_src.ncols} 列）")


def download_csrc_files(output_dir: str = "output") -> List[str]:
    """直接下载 CSRC 证券/基金/期货公司名录附件文件（不转 Excel）

    从 CSRC 页面找到附件链接（xlsx/xls），下载原始文件。

    Returns:
        下载的文件路径列表
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    sources = [
        ("证券公司名录", CSRC_SECURITIES_URL),
        ("基金管理公司名录", CSRC_FUND_URL),
        ("期货公司名录", CSRC_FUTURES_URL),
    ]

    files = []
    for label, page_url in sources:
        logger.info(f"正在获取 {label} 页面: {page_url}")
        html = _fetch_html(page_url)
        if not html:
            logger.warning(f"无法访问 {label} 页面")
            continue

        # 查找附件链接（xlsx/xls）
        attachment_url = _find_xlsx_url(html, page_url)
        if not attachment_url:
            logger.warning(f"{label} 页面未找到附件链接")
            continue

        logger.info(f"正在下载 {label}: {attachment_url}")
        data = _download_file(attachment_url, timeout=120)
        if not data:
            logger.warning(f"下载 {label} 失败")
            continue

        # 保留原始文件扩展名
        ext = Path(attachment_url.split("?")[0]).suffix or ".xlsx"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{label}_{timestamp}{ext}"
        filepath = Path(output_dir) / filename
        filepath.write_bytes(data)
        logger.info(f"{label} 已保存: {filepath} ({len(data)} bytes)")
        files.append(str(filepath))

    return files


def fetch_bank_insurance_list() -> Dict[str, List[Dict]]:
    """获取银行保险法人名单（通过 NFRA CDN PDF 直接下载）"""
    result = {"bank": [], "insurance": []}

    # ── 银行名单 ──
    logger.info("正在获取银行业金融机构法人名单...")
    pdf_bytes = _download_file(NFRA_BANK_PDF_URL_FALLBACK)
    if pdf_bytes:
        rows = _parse_pdf_table(pdf_bytes)
        for row in rows:
            if _is_header_row(row):
                continue
            if len(row) >= 2:
                result["bank"].append({
                    "name": row[1] if len(row) > 1 else row[0] if row else "",
                    "code": row[3] if len(row) > 3 else "",
                    "type": row[4] if len(row) > 4 else "",
                })
    logger.info(f"银行法人名单: {len(result['bank'])} 家")

    # ── 保险名单 ──
    logger.info("正在获取保险机构法人名单...")
    pdf_bytes = _download_file(NFRA_INSURANCE_PDF_URL_FALLBACK)
    if pdf_bytes:
        rows = _parse_pdf_table(pdf_bytes)
        for row in rows:
            if _is_header_row(row):
                continue
            if len(row) >= 2:
                result["insurance"].append({
                    "name": row[1] if len(row) > 1 else row[0] if row else "",
                    "code": row[3] if len(row) > 3 else "",
                    "type": row[4] if len(row) > 4 else "",
                })
    logger.info(f"保险法人名单: {len(result['insurance'])} 家")
    return result


def fetch_securities_fund_list() -> Dict[str, List[Dict]]:
    """获取证券基金期货公司名单"""
    result = {"securities": [], "funds": [], "futures": []}

    # 获取证券公司名录
    logger.info("正在获取证券公司名录...")
    html = _fetch_html(CSRC_SECURITIES_URL)
    if html:
        xlsx_url = _find_xlsx_url(html, CSRC_SECURITIES_URL)
        if xlsx_url:
            data = _download_file(xlsx_url)
            if data:
                rows = _parse_xlsx(data)
                for row in rows:
                    if _is_header_row(row, ["单位名称", "公司名称", "序号"]):
                        continue
                    if len(row) >= 2:
                        result["securities"].append({
                            "name": row[1] if len(row) > 1 else row[0] if row else "",
                            "addr": row[2] if len(row) > 2 else "",
                        })
        if not result["securities"]:
            tables = _parse_html_table(html)
            for chunk in _chunk_tables(tables):
                for row in chunk[1:]:
                    if len(row) >= 2 and not _is_header_row(row):
                        result["securities"].append({
                            "name": row[1] if len(row) > 1 else row[0] if row else "",
                            "addr": row[2] if len(row) > 2 else "",
                        })
    logger.info(f"证券公司名录: {len(result['securities'])} 家")

    # 获取基金公司名录
    logger.info("正在获取基金管理公司名录...")
    html = _fetch_html(CSRC_FUND_URL)
    if html:
        xlsx_url = _find_xlsx_url(html, CSRC_FUND_URL)
        if xlsx_url:
            data = _download_file(xlsx_url)
            if data:
                rows = _parse_xlsx(data)
                for row in rows:
                    if _is_header_row(row, ["公司名称", "单位名称", "序号"]):
                        continue
                    if len(row) >= 2:
                        result["funds"].append({
                            "name": row[1] if len(row) > 1 else row[0] if row else "",
                            "addr": row[2] if len(row) > 2 else "",
                        })
        if not result["funds"]:
            tables = _parse_html_table(html)
            for chunk in _chunk_tables(tables):
                for row in chunk[1:]:
                    if len(row) >= 2 and not _is_header_row(row):
                        result["funds"].append({
                            "name": row[1] if len(row) > 1 else row[0] if row else "",
                            "addr": row[2] if len(row) > 2 else "",
                        })
    logger.info(f"基金公司名录: {len(result['funds'])} 家")

    # 获取期货公司名录
    logger.info("正在获取期货公司名录...")
    html = _fetch_html(CSRC_FUTURES_URL)
    if html:
        xlsx_url = _find_xlsx_url(html, CSRC_FUTURES_URL)
        if xlsx_url:
            data = _download_file(xlsx_url)
            if data:
                rows = _parse_xlsx(data)
                for row in rows:
                    if _is_header_row(row, ["期货公司名称", "序号", "辖区"], name_col_index=2):
                        continue
                    if len(row) >= 2:
                        # 期货公司表格：序号 | 辖区 | 期货公司名称
                        name = row[2] if len(row) > 2 else row[1] if len(row) > 1 else ""
                        result["futures"].append({
                            "name": name,
                            "region": row[1] if len(row) > 1 else "",
                        })
        if not result["futures"]:
            tables = _parse_html_table(html)
            for row in tables[1:]:
                if not _is_header_row(row, ["期货公司名称", "序号", "辖区"], name_col_index=2):
                    if len(row) >= 2:
                        name = row[2] if len(row) > 2 else row[1] if len(row) > 1 else ""
                        region = row[1] if len(row) > 1 else ""
                        result["futures"].append({"name": name, "region": region})
    logger.info(f"期货公司名录: {len(result['futures'])} 家")
    return result


def fetch_all_institution_lists() -> Dict[str, List[Dict]]:
    """获取所有法人名单"""
    bank_insurance = fetch_bank_insurance_list()
    sec_fund = fetch_securities_fund_list()

    return {
        "bank": bank_insurance["bank"],
        "insurance": bank_insurance["insurance"],
        "securities": sec_fund["securities"],
        "funds": sec_fund["funds"],
        "futures": sec_fund["futures"],
    }


def write_institution_excel(
    data: Dict[str, List[Dict]],
    output_dir: str = "output",
    prefix: str = "金融机构法人名录",
) -> str:
    """将法人名单写入Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cf = Font(name="微软雅黑", size=10)
    align = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_sheet(ws, title, headers, rows_data):
        ws.title = title
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf
            c.fill = hfill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        for i, row in enumerate(rows_data, 1):
            for col, val in enumerate(row, 1):
                c = ws.cell(row=i + 1, column=col, value=val)
                c.font = cf
                c.alignment = align
                c.border = border

    # 根据提供的数据类型写入不同Sheet
    sheet_specs = [
        ("bank", "银行法人名单", ["序号", "机构名称", "机构编码", "机构类型"]),
        ("insurance", "保险法人名单", ["序号", "机构名称", "机构编码", "机构类型"]),
        ("securities", "证券公司名单", ["序号", "公司名称", "地址"]),
        ("funds", "基金公司名单", ["序号", "公司名称", "地址"]),
        ("futures", "期货公司名单", ["序号", "辖区", "期货公司名称"]),
    ]

    first = True
    for key, title, headers in sheet_specs:
        items = data.get(key, [])
        if not items:
            continue
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()

        rows_data = []
        for i, item in enumerate(items, 1):
            if key == "securities":
                rows_data.append([i, item.get("name", ""), item.get("addr", "")])
            elif key == "funds":
                rows_data.append([i, item.get("name", ""), item.get("addr", "")])
            elif key == "futures":
                rows_data.append([i, item.get("region", ""), item.get("name", "")])
            else:
                rows_data.append([i, item.get("name", ""), item.get("code", ""), item.get("type", "")])

        _write_sheet(ws, title, headers, rows_data)

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        if key in ("bank", "insurance"):
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 18
        elif key in ("securities", "funds"):
            ws.column_dimensions["C"].width = 50
        elif key == "futures":
            ws.column_dimensions["C"].width = 40

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    filepath = str(Path(output_dir) / filename)
    wb.save(filepath)
    logger.info(f"法人名录Excel已保存: {filepath}")
    return filepath