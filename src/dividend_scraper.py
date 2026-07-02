#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分红公告爬取模块
使用 AKShare 多数据源获取准确的 A股/H股 分红数据：
  - A股基础数据: stock_fhps_detail_em (东方财富F10) — 总股本、每股分红、股权登记日等
  - A股派息日:   stock_dividend_cninfo (巨潮资讯) — 准确的现金红利发放日
  - A股分红总额: stock_fhps_detail_ths (同花顺) — 准确的分红总额
  - H股派息日:   stock_hk_dividend_payout_em (东方财富港股) — 准确的发放日
"""

import re
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Callable, Tuple

import akshare as ak
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# 请求头（东方财富自有API）
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://data.eastmoney.com/",
    "Accept": "application/json, text/plain, */*",
}

# ========== 缓存 ==========
# A股/H股股本缓存: {code: (a_shares, h_shares, hk_code)}
_share_count_cache: Dict[str, Optional[Tuple[int, int, str]]] = {}
# HK代码缓存: {a_code: hk_code}
_hk_code_cache: Dict[str, str] = {}
# AKShare API结果缓存: {cache_key: DataFrame}
_api_cache: Dict[str, object] = {}


# ==================== 辅助函数 ====================

def _safe_float(value) -> float:
    """安全转换为浮点数"""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _normalize_date(date_str) -> str:
    """标准化日期格式为 YYYY-MM-DD"""
    if not date_str:
        return ""
    date_str = str(date_str).strip()
    # 处理 pandas NaT
    if date_str in ("NaT", "nan", "None", ""):
        return ""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str


def _report_type(report_date: str) -> str:
    """根据报告期判定报告类型"""
    if not report_date:
        return ""
    dt = _normalize_date(report_date)
    if dt.endswith("-12-31"):
        return "年报"
    elif dt.endswith("-06-30"):
        return "中报"
    elif dt.endswith("-03-31"):
        return "一季报"
    elif dt.endswith("-09-30"):
        return "三季报"
    return ""


def _report_year(report_date: str) -> str:
    """根据报告期提取报告年度"""
    if not report_date:
        return ""
    dt = _normalize_date(report_date)
    if dt:
        return dt[:4]
    return ""


def _period_to_date(period_str: str) -> str:
    """
    将中文报告期转换为 YYYY-MM-DD 格式
    例: "2025年报" → "2025-12-31", "2025中报" → "2025-06-30"
    """
    if not period_str:
        return ""
    period_str = str(period_str).strip()
    # 已经是 YYYY-MM-DD 格式
    if re.match(r'^\d{4}-\d{2}-\d{2}', period_str):
        return period_str[:10]
    # 中文格式: "2025年报", "2025中报", "2025一季报", "2025三季报"
    match = re.match(r'^(\d{4})(年报|中报|半年报|一季报|三季报)', period_str)
    if match:
        year = match.group(1)
        ptype = match.group(2)
        if ptype in ('年报',):
            return f"{year}-12-31"
        elif ptype in ('中报', '半年报'):
            return f"{year}-06-30"
        elif ptype == '一季报':
            return f"{year}-03-31"
        elif ptype == '三季报':
            return f"{year}-09-30"
    return period_str


def _parse_amount(amount_str: str) -> float:
    """解析金额字符串: "1.35亿" → 135000000, "5000万" → 50000000, "--" → 0"""
    if not amount_str or str(amount_str).strip() in ('--', '', 'nan', 'None'):
        return 0.0
    amount_str = str(amount_str).strip().replace(",", "")
    try:
        if '亿' in amount_str:
            return float(amount_str.replace('亿', '')) * 1e8
        elif '万' in amount_str:
            return float(amount_str.replace('万', '')) * 1e4
        else:
            return float(amount_str)
    except ValueError:
        return 0.0


# ==================== HK代码查找 ====================

def _lookup_hk_code(a_code: str) -> str:
    """
    通过A股代码查找对应的H股代码（东方财富F10 API）

    Returns:
        H股代码（纯数字字符串如 "02318"），未找到返回空串
    """
    if not a_code:
        return ""
    if a_code in _hk_code_cache:
        return _hk_code_cache[a_code]

    try:
        f10_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
        f10_params = {
            "reportName": "RPT_F10_ORG_BASICINFO",
            "columns": "STR_CODEH",
            "pageNumber": 1,
            "pageSize": 1,
            "source": "WEB",
            "client": "WEB",
            "filter": f'(SECURITY_CODE="{a_code}")',
        }
        resp = requests.get(f10_url, params=f10_params, headers=HEADERS, timeout=15)
        f10_data = resp.json()
        if f10_data.get("success") and f10_data.get("result") and f10_data["result"].get("data"):
            str_codeh = f10_data["result"]["data"][0].get("STR_CODEH", "")
            if str_codeh:
                # "02318.HK" → "02318"
                hk_code = str_codeh.split(".")[0].strip()
                if hk_code.isdigit():
                    _hk_code_cache[a_code] = hk_code
                    logger.info(f"  {a_code} → H股 {hk_code}")
                    return hk_code
    except Exception as e:
        logger.debug(f"查找 {a_code} H股代码失败: {e}")

    _hk_code_cache[a_code] = ""
    return ""


# ==================== A/H股本获取 ====================

def _get_ah_share_counts(code: str) -> Optional[Tuple[int, int]]:
    """
    获取A股/H股股本数量

    通过东方财富行情API获取A股股本，H股股本 = 总股本 - A股股本

    Returns:
        (a_shares, h_shares) 或 None（无H股）
    """
    if code in _share_count_cache:
        cached = _share_count_cache[code]
        if cached:
            return cached[0], cached[1]
        return None

    # 带重试的请求（东方财富API频率限制较严）
    max_retries = 3
    for attempt in range(max_retries):
        try:
            market_prefix = "0" if code.startswith("0") or code.startswith("3") else "1"
            secid = f"{market_prefix}.{code}"

            quote_url = "https://push2.eastmoney.com/api/qt/stock/get"
            params = {
                "secid": secid,
                "fields": "f84,f85",
                "ut": "fa5fd1943c7b386f172d6893dbbd4e1f",
            }
            resp = requests.get(quote_url, params=params, headers=HEADERS, timeout=15)
            data = resp.json()
            if not data.get("data"):
                _share_count_cache[code] = None
                return None

            total_shares = int(_safe_float(data["data"].get("f84", 0)))
            a_float = int(_safe_float(data["data"].get("f85", 0)))

            if total_shares <= 0:
                _share_count_cache[code] = None
                return None

            # 通过F10获取H股代码
            hk_code = _lookup_hk_code(code)
            if hk_code:
                a_shares = a_float
                h_shares = total_shares - a_shares
                if h_shares > 0:
                    _share_count_cache[code] = (a_shares, h_shares, hk_code)
                    logger.info(f"  {code}: A股={a_shares:,}, H股={h_shares:,} (HK:{hk_code})")
                    return (a_shares, h_shares)

            _share_count_cache[code] = None
            return None

        except Exception as e:
            if attempt < max_retries - 1:
                wait = (attempt + 1) * 1.5
                logger.debug(f"获取 {code} AH股本失败 (重试 {attempt + 1}/{max_retries}): {e}")
                time.sleep(wait)
            else:
                logger.debug(f"获取 {code} AH股本失败: {e}")
                _share_count_cache[code] = None
                return None

    _share_count_cache[code] = None
    return None


# ==================== A股分红数据获取（新多数据源方案） ====================

def _fetch_single_stock(code: str, start_year: int, end_year: int) -> List[Dict]:
    """
    获取单只股票的分红数据（多数据源合并）

    数据源:
      1. stock_fhps_detail_em  — 东方财富F10: 总股本、每股分红、股权登记日、除权除息日、方案进度
      2. stock_dividend_cninfo  — 巨潮资讯: 准确的派息日（现金红利发放日）
      3. stock_fhps_detail_ths — 同花顺: 准确的分红总额

    合并策略:
      - 派息日(PAYMENT_DATE): 巨潮资讯 > 东方财富除权除息日
      - 分红总额(TOTAL_AMOUNT): 同花顺 > 东方财富计算值
      - 公告日(NOTICE_DATE): 巨潮资讯方案公告日 > 东方财富最新公告日
    """
    records = []

    # 1. 获取东方财富基础数据
    try:
        em_df = ak.stock_fhps_detail_em(symbol=code)
    except Exception as e:
        logger.warning(f"  {code} stock_fhps_detail_em 失败: {e}")
        return records

    if em_df is None or em_df.empty:
        return records

    # 2. 获取巨潮资讯派息日数据
    cninfo_map = {}  # {报告期: row}
    try:
        cninfo_df = ak.stock_dividend_cninfo(symbol=code)
        if cninfo_df is not None and not cninfo_df.empty:
            for _, row in cninfo_df.iterrows():
                period = _period_to_date(row.get("报告时间", ""))
                if period:
                    cninfo_map[period] = row
    except Exception as e:
        logger.debug(f"  {code} stock_dividend_cninfo 失败: {e}")

    # 3. 获取同花顺分红总额数据
    ths_map = {}  # {报告期: row}
    try:
        ths_df = ak.stock_fhps_detail_ths(symbol=code)
        if ths_df is not None and not ths_df.empty:
            for _, row in ths_df.iterrows():
                period = _period_to_date(row.get("报告期", ""))
                if period:
                    ths_map[period] = row
    except Exception as e:
        logger.debug(f"  {code} stock_fhps_detail_ths 失败: {e}")

    # 4. 遍历东方财富记录，合并各数据源
    for _, em_row in em_df.iterrows():
        # 方案进度筛选：只保留"实施分配"或"实施方案"
        plan_status = str(em_row.get("方案进度", ""))
        if "实施" not in plan_status:
            continue

        # 报告期
        report_period = str(em_row.get("报告期", ""))[:10]
        if not report_period or report_period == "NaT":
            continue

        # 按年份筛选
        try:
            year = int(report_period[:4])
        except ValueError:
            continue
        if year < start_year or year > end_year:
            continue

        # === 合并巨潮资讯数据 ===
        cninfo_row = cninfo_map.get(report_period)
        if cninfo_row is not None:
            # 派息日：巨潮资讯的"派息日"字段是准确的现金发放日
            payment_date = _normalize_date(cninfo_row.get("派息日", ""))
            # 公告日：巨潮资讯的"实施方案公告日期"
            notice_date = _normalize_date(cninfo_row.get("实施方案公告日期", ""))
        else:
            # 回退：用除权除息日作为派息日（旧逻辑）
            payment_date = _normalize_date(em_row.get("除权除息日", ""))
            notice_date = _normalize_date(em_row.get("最新公告日期", ""))

        # === 合并同花顺数据 ===
        ths_row = ths_map.get(report_period)
        if ths_row is not None:
            total_amount = _parse_amount(ths_row.get("分红总额", ""))
        else:
            total_amount = 0.0

        # === 从东方财富计算基础值 ===
        # 现金分红比例（每10股）
        bonus_ratio = _safe_float(em_row.get("现金分红-现金分红比例", 0))
        per_share = bonus_ratio / 10.0 if bonus_ratio > 0 else 0

        # 总股本
        total_shares = int(_safe_float(em_row.get("总股本", 0)))

        # 如果同花顺没有分红总额，从东方财富计算
        if total_amount <= 0 and total_shares > 0 and per_share > 0:
            total_amount = round(per_share * total_shares, 2)

        # 股权登记日
        regist_date = _normalize_date(em_row.get("股权登记日", ""))

        # 除权除息日
        ex_dividend_date = _normalize_date(em_row.get("除权除息日", ""))

        record = {
            "SECURITY_CODE": code,
            "SECURITY_NAME": str(em_row.get("业绩披露日期", ""))[:0],  # 后面由app.py补
            "PLAN_NOTICE_DATE": notice_date,
            "NOTICE_DATE": notice_date,
            "REGIST_DATE": regist_date,
            "EX_DIVIDEND_DATE": ex_dividend_date,
            "PAYMENT_DATE": payment_date,  # 优先使用巨潮资讯的派息日
            "H_PAYMENT_DATE": "",  # 后续由H股数据补充
            "CASH_DIVIDEND_PER_SHARE": round(per_share, 4),
            "TOTAL_AMOUNT": round(total_amount, 2),
            "TOTAL_SHARES": total_shares,
            "REPORT_TYPE": _report_type(report_period),
            "REPORT_YEAR": _report_year(report_period),
            "PLAN_STATUS": plan_status,
            "REPORT_PERIOD": report_period,
            "IMPL_PLAN_PROFILE": str(em_row.get("现金分红-现金分红比例描述", "")),
        }

        records.append(record)

    return records


# ==================== H股分红数据获取 ====================

def _fetch_hk_dividend_dates(a_code: str) -> Dict[str, str]:
    """
    获取H股红利派发日期（使用 stock_hk_dividend_payout_em）

    流程:
      1. 通过东方财富F10 API查找H股代码
      2. 调用 stock_hk_dividend_payout_em 获取分红派息列表
      3. 按财政年度匹配发放日

    Returns: {报告期: 发放日, ...}  如 {"2025-12-31": "2026-07-15"}
    """
    if not a_code:
        return {}

    hk_code = _lookup_hk_code(a_code)
    if not hk_code:
        return {}

    try:
        df = ak.stock_hk_dividend_payout_em(symbol=hk_code)
    except Exception as e:
        logger.warning(f"  获取H股分红日期失败 ({a_code} → HK{hk_code}): {e}")
        return {}

    if df is None or df.empty:
        return {}

    result = {}
    for _, row in df.iterrows():
        fiscal_year = str(row.get("财政年度", ""))
        payment_date = _normalize_date(row.get("发放日", ""))
        dist_type = str(row.get("分配类型", ""))
        if fiscal_year and payment_date:
            if "中期" in dist_type:
                period = f"{fiscal_year}-06-30"
            else:
                period = f"{fiscal_year}-12-31"
            result[period] = payment_date

    logger.info(f"  {a_code} → H股 {hk_code}: {len(result)} 条派息日")
    return result


def _fetch_hk_dividend_list(a_code: str, start_year: int = 0, end_year: int = 9999) -> List[Dict]:
    """
    获取H股分红公告列表（使用 stock_hk_dividend_payout_em）

    Returns:
        [{ "FiscalYear": "2025", "YearEnd": "2025-12-31",
           "Scheme": "每股派人民币1.75元", "DistributionType": "年度分配",
           "ExDividendDate": "2026-06-02", "RecordDate": "2026/06/04-2026/06/09",
           "PaymentDate": "2026-07-15", "AnnouncementDate": "2026-06-10",
           "Status": "实施" }, ...]
    """
    if not a_code:
        return []

    hk_code = _lookup_hk_code(a_code)
    if not hk_code:
        return []

    try:
        df = ak.stock_hk_dividend_payout_em(symbol=hk_code)
    except Exception as e:
        logger.warning(f"  获取H股分红公告列表失败 ({a_code} → HK{hk_code}): {e}")
        return []

    if df is None or df.empty:
        return []

    result = []
    for _, row in df.iterrows():
        fiscal_year = str(row.get("财政年度", ""))
        dist_type = str(row.get("分配类型", ""))

        # 按年份范围筛选
        try:
            fy = int(fiscal_year)
        except ValueError:
            continue
        if fy < start_year or fy > end_year:
            continue

        # 根据分配类型确定YearEnd
        if "中期" in dist_type:
            year_end = f"{fiscal_year}-06-30" if fiscal_year else ""
        else:
            year_end = f"{fiscal_year}-12-31" if fiscal_year else ""

        result.append({
            "FiscalYear": fiscal_year,
            "YearEnd": year_end,
            "Scheme": str(row.get("分红方案", "")),
            "DistributionType": dist_type,
            "ExDividendDate": _normalize_date(row.get("除净日", "")),
            "RecordDate": str(row.get("截至过户日", "")),
            "PaymentDate": _normalize_date(row.get("发放日", "")),
            "AnnouncementDate": _normalize_date(row.get("最新公告日期", "")),
            "Status": "实施",
        })

    logger.info(f"  {a_code} → H股 {hk_code}: {len(result)} 条分红公告")
    return result


# ==================== 主入口 ====================

def fetch_dividend_data(
    stock_codes: List[str],
    start_year: int,
    end_year: int,
    progress_callback: Optional[Callable] = None,
) -> Tuple[List[Dict], List[Tuple[str, str, List[Dict]]], List[str]]:
    """
    获取指定股票代码列表的分红公告数据

    Args:
        stock_codes: 股票代码列表，如 ['000001', '600036']
        start_year: 起始年份
        end_year: 结束年份
        progress_callback: 进度回调 (current, total, message)

    Returns:
        (分红记录列表, H股分红公告列表, 未找到数据的股票代码列表)
    """
    all_records = []
    not_found = []
    total = len(stock_codes)

    # ===== 阶段1: 获取A股分红数据 =====
    for idx, code in enumerate(stock_codes):
        if progress_callback:
            progress_callback(idx + 1, total, f"正在查询 {code} 分红数据...")

        records = _fetch_single_stock(code, start_year, end_year)
        if records:
            all_records.extend(records)
            logger.info(f"  {code}: 获取到 {len(records)} 条分红记录")
        else:
            not_found.append(code)
            logger.warning(f"  {code}: 未获取到分红数据")

        time.sleep(0.3)  # 请求间隔

    # ===== 阶段2: 补充公司名称（从东方财富行情API，去重） =====
    if progress_callback:
        progress_callback(0, 0, "正在补充公司名称...")

    name_cache: Dict[str, str] = {}
    for record in all_records:
        code = record["SECURITY_CODE"]
        if code not in name_cache:
            name_cache[code] = _get_company_name(code)
            time.sleep(0.15)
        if name_cache[code]:
            record["SECURITY_NAME"] = name_cache[code]

    # ===== 阶段3: 计算A股/H股分配金额 =====
    if progress_callback:
        progress_callback(0, 0, "正在计算A股/H股分配金额...")

    time.sleep(0.5)  # 避免东方财富API频率限制

    for record in all_records:
        code = record["SECURITY_CODE"]
        per_share = record["CASH_DIVIDEND_PER_SHARE"]
        total_shares = record.get("TOTAL_SHARES", 0)

        ah_counts = _get_ah_share_counts(code)
        if ah_counts:
            a_shares, h_shares = ah_counts
            record["A_SHARE_AMOUNT"] = round(per_share * a_shares, 2)
            record["H_SHARE_AMOUNT"] = round(per_share * h_shares, 2)
        else:
            # 纯A股公司：A股分配金额 = 总分配金额，H股分配金额 = 0
            record["A_SHARE_AMOUNT"] = record["TOTAL_AMOUNT"]
            record["H_SHARE_AMOUNT"] = 0

    # ===== 阶段4: 获取H股红利派发日期 =====
    if progress_callback:
        progress_callback(0, 0, "正在获取H股红利派发日期...")

    a_codes = list(set(r["SECURITY_CODE"] for r in all_records if _get_ah_share_counts(r["SECURITY_CODE"])))
    hk_dates_cache: Dict[str, Dict[str, str]] = {}

    for a_code in a_codes:
        hk_dates_cache[a_code] = _fetch_hk_dividend_dates(a_code)
        if hk_dates_cache[a_code]:
            logger.info(f"  {a_code}: 获取到 {len(hk_dates_cache[a_code])} 条派息日")
        else:
            logger.info(f"  {a_code}: 未获取到H股派息日")
        time.sleep(0.3)

    # 填充分红记录的H股派息日（按年份匹配）
    for record in all_records:
        a_code = record["SECURITY_CODE"]
        if a_code in hk_dates_cache:
            report_period = record.get("REPORT_PERIOD", "")
            if report_period and report_period in hk_dates_cache[a_code]:
                record["H_PAYMENT_DATE"] = hk_dates_cache[a_code][report_period]

    # ===== 阶段5: 获取H股分红公告列表 =====
    hk_list_data: List[Tuple[str, str, List[Dict]]] = []
    if progress_callback:
        progress_callback(0, 0, "正在获取H股分红公告...")

    for a_code in a_codes:
        hk_list = _fetch_hk_dividend_list(a_code, start_year, end_year)
        if hk_list:
            company_name = _get_company_name(a_code)
            hk_list_data.append((a_code, company_name, hk_list))
        time.sleep(0.3)

    return all_records, hk_list_data, not_found


def _get_company_name(code: str) -> str:
    """获取公司名称（东方财富行情API）"""
    try:
        market_prefix = "0" if code.startswith("0") or code.startswith("3") else "1"
        secid = f"{market_prefix}.{code}"
        quote_url = "https://push2.eastmoney.com/api/qt/stock/get"
        params = {
            "secid": secid,
            "fields": "f58",
            "ut": "fa5fd1943c7b386f172d6893dbbd4e1f",
        }
        resp = requests.get(quote_url, params=params, headers=HEADERS, timeout=10)
        data = resp.json()
        if data.get("data"):
            return str(data["data"].get("f58", ""))
    except Exception:
        pass
    return ""


# ==================== Excel输出 ====================

def write_dividend_excel(
        records: List[Dict],
        start_year: int,
        end_year: int,
        output_dir: str = "output",
        hk_records: Optional[List[Tuple[str, str, List[Dict]]]] = None,
    ) -> str:
    """
    将分红数据写入Excel文件

    输出列:
      - 股票代码、股票名称、公告日、股权登记日、现金红利发放日、
        每股分红(元)、分配金额(元)、A股分配金额、H股分配金额
    """
    from collections import OrderedDict

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{start_year}-{end_year}年分红公告"

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    group_font = Font(name="微软雅黑", size=10, bold=True, color="8E44AD")
    group_fill = PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid")

    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )

    # 标题行
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"{start_year}-{end_year}年 上市公司分红公告"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="8E44AD")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 表头
    headers = [
        "股票代码", "股票名称", "公告日", "报告年度", "报告类型",
        "股权登记日", "现金红利发放日", "H股红利派发日期",
        "每股分红(元)", "分配金额(元)", "A股分配金额(元)", "H股分配金额(元)",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 按股票代码分组排序
    groups = OrderedDict()
    for rec in records:
        code = rec.get("SECURITY_CODE", "未知")
        if code not in groups:
            groups[code] = []
        groups[code].append(rec)

    current_row = 3
    for code, recs in groups.items():
        # 按公告日降序排列
        recs.sort(key=lambda x: x.get("PLAN_NOTICE_DATE", ""), reverse=True)

        name = recs[0].get("SECURITY_NAME", "") if recs else ""

        # 分组标题行
        ws.merge_cells(f"A{current_row}:L{current_row}")
        group_cell = ws.cell(row=current_row, column=1, value=f"{code}  {name}  ({len(recs)}条)")
        group_cell.font = group_font
        group_cell.fill = group_fill
        for c in range(1, 13):
            ws.cell(row=current_row, column=c).border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for rec in recs:
            ws.cell(row=current_row, column=1, value=rec.get("SECURITY_CODE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value=rec.get("SECURITY_NAME", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=2).border = thin_border

            ws.cell(row=current_row, column=3, value=rec.get("NOTICE_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=3).border = thin_border

            ws.cell(row=current_row, column=4, value=rec.get("REPORT_YEAR", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=4).border = thin_border

            ws.cell(row=current_row, column=5, value=rec.get("REPORT_TYPE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=5).border = thin_border

            ws.cell(row=current_row, column=6, value=rec.get("REGIST_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=6).border = thin_border

            ws.cell(row=current_row, column=7, value=rec.get("PAYMENT_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=7).border = thin_border

            ws.cell(row=current_row, column=8, value=rec.get("H_PAYMENT_DATE", "")).alignment = cell_alignment
            ws.cell(row=current_row, column=8).border = thin_border

            ws.cell(row=current_row, column=9, value=rec.get("CASH_DIVIDEND_PER_SHARE", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=9).number_format = "#,##0.0000"
            ws.cell(row=current_row, column=9).border = thin_border

            ws.cell(row=current_row, column=10, value=rec.get("TOTAL_AMOUNT", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=10).number_format = "#,##0.00"
            ws.cell(row=current_row, column=10).border = thin_border

            ws.cell(row=current_row, column=11, value=rec.get("A_SHARE_AMOUNT", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=11).number_format = "#,##0.00"
            ws.cell(row=current_row, column=11).border = thin_border

            ws.cell(row=current_row, column=12, value=rec.get("H_SHARE_AMOUNT", 0)).alignment = cell_alignment
            ws.cell(row=current_row, column=12).number_format = "#,##0.00"
            ws.cell(row=current_row, column=12).border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 20

    # 冻结
    ws.freeze_panes = "A3"

    # ========== H股分红公告 Sheet ==========
    if hk_records:
        ws_hk = wb.create_sheet("H股分红公告")

        # 标题行
        ws_hk.merge_cells("A1:H1")
        title_hk = ws_hk["A1"]
        title_hk.value = "H股分红公告"
        title_hk.font = Font(name="微软雅黑", size=14, bold=True, color="8E44AD")
        title_hk.alignment = Alignment(horizontal="center", vertical="center")
        ws_hk.row_dimensions[1].height = 36

        # 表头
        hk_headers = ["股票代码", "股票名称", "公告日", "财政年度", "分配类型", "分红方案", "除净日", "发放日"]
        for col_idx, header in enumerate(hk_headers, 1):
            cell = ws_hk.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws_hk.row_dimensions[2].height = 28

        current_row = 3
        for code, name, recs in hk_records:
            if not recs:
                continue

            # 分组标题
            ws_hk.merge_cells(f"A{current_row}:H{current_row}")
            group_cell = ws_hk.cell(row=current_row, column=1, value=f"{code}  {name}  ({len(recs)}条)")
            group_cell.font = group_font
            group_cell.fill = group_fill
            for c in range(1, 9):
                ws_hk.cell(row=current_row, column=c).border = thin_border
            ws_hk.row_dimensions[current_row].height = 26
            current_row += 1

            for rec in recs:
                ws_hk.cell(row=current_row, column=1, value=code).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=1).border = thin_border
                ws_hk.cell(row=current_row, column=2, value=name).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=2).border = thin_border
                ws_hk.cell(row=current_row, column=3, value=rec.get("AnnouncementDate", "")).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=3).border = thin_border
                ws_hk.cell(row=current_row, column=4, value=rec.get("FiscalYear", "")).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=4).border = thin_border
                ws_hk.cell(row=current_row, column=5, value=rec.get("DistributionType", "")).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=5).border = thin_border
                ws_hk.cell(row=current_row, column=6, value=rec.get("Scheme", "")).alignment = Alignment(wrap_text=True, vertical="center")
                ws_hk.cell(row=current_row, column=6).border = thin_border
                ws_hk.cell(row=current_row, column=7, value=rec.get("ExDividendDate", "")).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=7).border = thin_border
                ws_hk.cell(row=current_row, column=8, value=rec.get("PaymentDate", "")).alignment = cell_alignment
                ws_hk.cell(row=current_row, column=8).border = thin_border
                ws_hk.row_dimensions[current_row].height = 30
                current_row += 1

        # H股sheet列宽
        ws_hk.column_dimensions["A"].width = 12
        ws_hk.column_dimensions["B"].width = 22
        ws_hk.column_dimensions["C"].width = 14
        ws_hk.column_dimensions["D"].width = 10
        ws_hk.column_dimensions["E"].width = 12
        ws_hk.column_dimensions["F"].width = 45
        ws_hk.column_dimensions["G"].width = 14
        ws_hk.column_dimensions["H"].width = 14
        ws_hk.freeze_panes = "A3"

    # 保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"分红公告_{start_year}-{end_year}年_{timestamp}.xlsx"
    filepath = str(Path(output_dir) / filename)
    wb.save(filepath)
    logger.info(f"分红Excel已保存: {filepath} (共{len(records)}条记录)")

    return filepath


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    def cb(cur, total, msg):
        print(f"\r[{cur}/{total}] {msg}", end="")

    records, hk_list, nf = fetch_dividend_data(["000001", "600036"], 2024, 2025, progress_callback=cb)
    print(f"\n获取 {len(records)} 条记录, 未找到: {nf}")
    if records:
        path = write_dividend_excel(records, 2024, 2025, hk_records=hk_list)
        print(f"已输出: {path}")