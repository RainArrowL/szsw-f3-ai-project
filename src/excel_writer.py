#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel写入模块
- 一企业一文件
- 每年度每个财务报表一个sheet
- 列名均为中文
"""

import logging
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Set, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from config import config

logger = logging.getLogger(__name__)


# Excel样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 数值列标识（根据字段名判断）
NUMERIC_KEYWORDS = [
    "金额", "合计", "总计", "净额", "净收入", "净支出", "净利润",
    "总额", "余额", "小计", "收益", "损益", "支出", "收入",
    "成本", "费用", "利润", "资产", "负债", "权益", "税",
    "现金", "借款", "股本", "资本", "盈余", "股", "利息",
    "工资", "每股", "折旧", "摊销",
]

# 人民币数字格式
CNY_FORMAT = '#,##0.00'


def _is_numeric_column(header: str) -> bool:
    """判断列是否为数值类型"""
    for kw in NUMERIC_KEYWORDS:
        if kw in header:
            return True
    return False


def _format_sheet(ws, headers: List[str], num_cols: int, num_rows: int):
    """
    美化Excel工作表格式

    参数:
        ws: 工作表对象
        headers: 表头列表
        num_cols: 列数
        num_rows: 数据行数（不含表头）
    """
    # 设置列宽
    for col_idx in range(1, num_cols + 1):
        header_text = headers[col_idx - 1] if col_idx <= len(headers) else ""
        if len(header_text) > 10:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(header_text) * 2.2)
        elif _is_numeric_column(header_text):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # 设置表头样式
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 设置数据行样式
    for row_idx in range(2, num_rows + 2):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            header = headers[col_idx - 1] if col_idx <= len(headers) else ""
            if _is_numeric_column(header):
                cell.alignment = NUMBER_ALIGNMENT
                if isinstance(cell.value, (int, float)):
                    cell.number_format = CNY_FORMAT
            else:
                cell.alignment = DATA_ALIGNMENT
                cell.font = DATA_FONT

    # 冻结首行
    ws.freeze_panes = "A2"
    # 添加自动筛选
    if num_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"


def _extract_year_from_date(date_str: str) -> str:
    """从日期字符串中提取年份"""
    if not date_str:
        return "未知年份"
    # 处理多种日期格式
    date_str = str(date_str)
    for sep in ["-", "/", "."]:
        if sep in date_str:
            return date_str.split(sep)[0]
    if len(date_str) >= 4:
        return date_str[:4]
    return date_str


def _write_notes_sheet(wb, records: List[Dict[str, Any]]):
    """写入附注sheet（所有年度合并）"""
    ws = wb.create_sheet(title="附注")

    # 收集所有非空字段名
    all_fields = []
    seen = set()
    # 按报告年度降序排列
    records = sorted(records, key=lambda x: str(x.get("报告年度", "")), reverse=True)

    for rec in records:
        for key in rec:
            if key not in seen and key not in SKIP_META_FIELDS:
                all_fields.append(key)
                seen.add(key)

    if not all_fields:
        ws.cell(row=1, column=1, value="暂无附注数据")
        return

    # 表头
    headers = ["报告年度", "报告类型"] + all_fields
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, rec in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=rec.get("报告年度", ""))
        ws.cell(row=row_idx, column=2, value=rec.get("报告类型", ""))
        for col_idx, field in enumerate(all_fields, 3):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(field, ""))

    _format_sheet(ws, headers, len(headers), len(records))


def write_company_excel(
    company_name: str,
    financial_data: Dict[str, List[Dict[str, Any]]],
    output_dir: str = None,
) -> str:
    """
    将单个公司的财务数据写入Excel

    参数:
        company_name: 公司名称（含股票代码）
        financial_data: 财务数据，格式:
            {"资产负债表": [...], "利润表": [...], "现金流量表": [...]}
        output_dir: 输出目录

    返回:
        生成的文件路径
    """
    if output_dir is None:
        output_dir = config.output_dir

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # 清理文件名中的非法字符
    safe_name = company_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    filepath = str(Path(output_dir) / f"{safe_name}_年报财务数据.xlsx")

    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 按报表类型分组数据，再将每个报表按年度分sheet
    for report_name, records in financial_data.items():
        if not records:
            # 创建空sheet
            ws = wb.create_sheet(title=f"{report_name}(无数据)")
            ws.cell(row=1, column=1, value="暂无数据")
            continue

        # 附注：单独一个sheet，所有年度在一起
        if report_name == "附注":
            _write_notes_sheet(wb, records)
            continue

        # 按年份分组
        year_groups: Dict[str, list] = {}
        for record in records:
            report_date = record.get("报告期", "") or record.get("报告日期", "")
            year = _extract_year_from_date(report_date)
            if year not in year_groups:
                year_groups[year] = []
            year_groups[year].append(record)

        # 为每年创建一个sheet
        for year in sorted(year_groups.keys(), reverse=True):
            year_records = year_groups[year]

            # Sheet名称限制31字符
            sheet_title = f"{report_name}_{year}年"
            if len(sheet_title) > 31:
                sheet_title = f"{year}年"

            # 如果同一年同一类型数据多行（如修正公告），只取最新一行
            if len(year_records) > 1:
                # 优先取年报（通常报告期是12-31）
                annual_records = [
                    r for r in year_records
                    if "12-31" in str(r.get("报告期", ""))
                    or "12-31" in str(r.get("报告日期", ""))
                ]
                if annual_records:
                    year_records = annual_records[:1]
                else:
                    year_records = year_records[:1]

            ws = wb.create_sheet(title=sheet_title)

            # 收集该报表所有字段（去重）
            headers = []
            seen = set()
            for record in year_records:
                for key in record:
                    if key not in seen:
                        headers.append(key)
                        seen.add(key)

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, record in enumerate(year_records, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 格式化
            _format_sheet(ws, headers, len(headers), len(year_records))

    wb.save(filepath)
    logger.info(f"Excel已保存: {filepath}")
    return filepath


def write_merged_by_report_type(
    all_company_data: Dict[str, Dict[str, List[Dict[str, Any]]]],
    company_industries: Dict[str, str],  # {code: industry_name}
    output_dir: str = None,
) -> List[str]:
    """
    按报表类型合并输出Excel：资产负债表、利润表、现金流量表各一个文件。
    若不同行业报表字段不同，按行业再拆分。

    每个Excel只有一个sheet，列为：公司名称 | 年份 | [财务字段...]

    参数:
        all_company_data: {公司名: {"资产负债表": [...], "利润表": [...], "现金流量表": [...]}}
        company_industries: {股票代码: 行业名称}  用于按行业分组
        output_dir: 输出目录

    返回:
        生成的文件路径列表
    """
    if output_dir is None:
        output_dir = config.output_dir
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    files = []

    # 报表类型 → 中文名
    report_types = ["资产负债表", "利润表", "现金流量表"]

    for report_name in report_types:
        # 收集该报表类型所有公司的数据，按行业分组
        # industry -> [(company_name, code, year, record), ...]
        industry_groups: Dict[str, List[tuple]] = defaultdict(list)
        # 记录每个行业的字段集合，用于判断是否需要拆分
        industry_fields: Dict[str, Set[str]] = defaultdict(set)

        for company_name, data in all_company_data.items():
            records = data.get(report_name, [])
            if not records:
                continue

            # 提取股票代码
            code = ""
            if "(" in company_name and ")" in company_name:
                code = company_name.split("(")[-1].split(")")[0]

            ind_name = company_industries.get(code, "其他")

            for record in records:
                report_date = record.get("报告期", "") or record.get("报告日期", "")
                year = _extract_year_from_date(report_date)

                industry_groups[ind_name].append((company_name, code, year, record))
                # 收集字段
                for key in record:
                    if key not in SKIP_META_FIELDS:
                        industry_fields[ind_name].add(key)

        # 判断是否需要按行业拆分：比较各行业的字段集合
        # 简单策略：如果只有一个行业或所有行业字段相同，合并为一个文件
        if not industry_groups:
            continue

        if len(industry_groups) == 1:
            # 只有一个行业，直接输出
            _write_single_merged_excel(
                report_name, list(industry_groups.values())[0],
                output_dir, files
            )
        else:
            # 多行业：检查字段差异
            field_sets = list(industry_fields.values())
            all_same = all(s == field_sets[0] for s in field_sets[1:])

            if all_same:
                # 字段相同，合并所有行业到一个文件
                all_rows = []
                for rows in industry_groups.values():
                    all_rows.extend(rows)
                _write_single_merged_excel(report_name, all_rows, output_dir, files)
            else:
                # 字段不同，按行业分别输出
                for ind_name, rows in industry_groups.items():
                    safe_ind = ind_name.replace("/", "_").replace("\\", "_").replace(":", "_")
                    _write_single_merged_excel(
                        f"{report_name}_{safe_ind}", rows, output_dir, files
                    )

    # 合并附注
    _write_merged_notes_excel(all_company_data, output_dir, files)

    return files


# 合并时跳过的元数据字段
SKIP_META_FIELDS = {
    "SECUCODE", "SECURITY_CODE", "SECURITY_NAME_ABBR", "ORG_CODE",
    "SECURITY_TYPE_CODE", "TRADE_MARKET_CODE", "DATE_TYPE_CODE",
    "REPORT_TYPE_CODE", "DATA_STATE", "MARKET", "REPORT_DATE",
    "NOTICE_DATE", "INDUSTRY_CODE", "INDUSTRY_NAME",
    "证券代码", "股票代码", "证券简称", "机构代码", "行业代码", "行业名称",
    "交易市场", "证券类型代码", "交易市场代码", "日期类型代码",
    "报告类型代码", "数据状态", "公告日期",
}


def _write_single_merged_excel(
    report_name: str,
    rows: List[tuple],  # [(company_name, code, year, record), ...]
    output_dir: str,
    files: List[str],
):
    """写入单个合并报表Excel文件"""
    from datetime import datetime

    # 按公司名称排序，同年份放在一起
    rows.sort(key=lambda x: (x[2], x[0]), reverse=True)  # 按年份降序，同公司相邻

    # 收集所有字段（联合所有记录）
    all_fields = []
    seen_fields = set()
    for _, _, _, record in rows:
        for key in record:
            if key not in seen_fields and key not in SKIP_META_FIELDS:
                all_fields.append(key)
                seen_fields.add(key)

    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]

    # 表头：公司名称 | 年份 | 字段1 | 字段2 | ...
    headers = ["公司名称", "年份"] + all_fields
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 写入数据
    for row_idx, (company_name, code, year, record) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=company_name)
        ws.cell(row=row_idx, column=2, value=year)

        for col_idx, field in enumerate(all_fields, 3):
            value = record.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 格式化
    _format_sheet(ws, headers, len(headers), len(rows))

    # 安全文件名
    safe_name = report_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = str(Path(output_dir) / f"{safe_name}_合并_{timestamp}.xlsx")
    wb.save(filepath)
    files.append(filepath)
    logger.info(f"合并Excel已保存: {filepath} ({len(rows)}行, {len(all_fields)}字段)")


def _write_merged_notes_excel(
    all_company_data: Dict[str, Dict[str, List[Dict[str, Any]]]],
    output_dir: str,
    files: List[str],
):
    """写入合并的附注Excel文件"""
    from datetime import datetime

    all_rows = []
    all_fields = []
    seen_fields = set()

    for company_name, data in all_company_data.items():
        notes = data.get("附注", [])
        if not notes:
            continue
        for rec in notes:
            all_rows.append((company_name, rec))
            for key in rec:
                if key not in seen_fields and key not in SKIP_META_FIELDS:
                    all_fields.append(key)
                    seen_fields.add(key)

    if not all_rows:
        return

    # 按公司名和报告年度排序
    all_rows.sort(key=lambda x: (x[0], str(x[1].get("报告年度", "0"))), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "附注"

    headers = ["公司名称", "报告年度", "报告类型"] + all_fields
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, (company_name, rec) in enumerate(all_rows, 2):
        ws.cell(row=row_idx, column=1, value=company_name)
        ws.cell(row=row_idx, column=2, value=rec.get("报告年度", ""))
        ws.cell(row=row_idx, column=3, value=rec.get("报告类型", ""))
        for col_idx, field in enumerate(all_fields, 4):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(field, ""))

    _format_sheet(ws, headers, len(headers), len(all_rows))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = str(Path(output_dir) / f"附注_合并_{timestamp}.xlsx")
    wb.save(filepath)
    files.append(filepath)
    logger.info(f"合并附注Excel已保存: {filepath} ({len(all_rows)}行, {len(all_fields)}字段)")


def write_industry_avg_excel(
    industry_name: str,
    industry_avg: Dict[str, Dict[str, List[Dict[str, Any]]]],
    output_dir: str = None,
) -> str:
    """
    将行业平均值写入单独的Excel文件

    参数：
        industry_name: 行业名称
        industry_avg: 行业均值数据 {"行业均值_资产负债表": {"2023": [...]}, ...}
        output_dir: 输出目录

    返回：
        文件路径
    """
    if output_dir is None:
        output_dir = config.output_dir

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    safe_name = industry_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    filepath = str(Path(output_dir) / f"行业均值_{safe_name}_年报财务数据.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for ia_label, ia_data in industry_avg.items():
        if not ia_data:
            # 空sheet
            ws = wb.create_sheet(title=f"{ia_label}(无数据)")
            ws.cell(row=1, column=1, value="暂无数据")
            continue

        for year in sorted(ia_data.keys(), reverse=True):
            year_records = ia_data[year]
            if not year_records:
                continue

            sheet_title = f"{ia_label}_{year}年"
            if len(sheet_title) > 31:
                # 截短：行业均值_BS_2023年，保留关键信息
                short_label = ia_label.replace("行业均值_", "均值_").replace("资产负债表", "BS").replace("利润表", "PL").replace("现金流量表", "CF")
                sheet_title = f"{short_label}_{year}年"[:31]

            ws = wb.create_sheet(title=sheet_title)

            headers = []
            seen = set()
            for record in year_records:
                for key in record:
                    if key not in seen:
                        headers.append(key)
                        seen.add(key)

            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            for row_idx, record in enumerate(year_records, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))

            _format_sheet(ws, headers, len(headers), len(year_records))

    wb.save(filepath)
    logger.info(f"行业均值Excel已保存: {filepath}")
    return filepath


def write_all_companies(
    all_data: Dict[str, Dict[str, List[Dict[str, Any]]]],
    output_dir: str = None,
    progress_callback=None,
    industry_avgs_map: Dict[str, Dict] = None,
) -> List[str]:
    """
    批量写入所有公司的Excel文件

    参数:
        all_data: 所有公司的财务数据
        output_dir: 输出目录
        progress_callback: 进度回调
        industry_avgs_map: 行业平均值映射 {公司名: 行业平均值数据}

    返回:
        生成的文件路径列表
    """
    if output_dir is None:
        output_dir = config.output_dir

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    files = []

    total = len(all_data)
    for idx, (company_name, data) in enumerate(all_data.items()):
        logger.info(f"正在写入 {idx + 1}/{total}: {company_name}")
        if progress_callback:
            progress_callback(idx + 1, total, f"正在写入: {company_name}")

        filepath = write_company_excel(company_name, data, output_dir)
        files.append(filepath)

    return files